#!/usr/bin/env python3
"""
Generate an Excel review sheet with all chunk_sim special rules.
Matches the format of docs/special_rules_review_matrix.md
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_rules_sheet():
    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    category_font = Font(bold=True, size=11)
    category_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    implemented_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    not_impl_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    phase_yes_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # =============================================
    # SHEET 1: Complete Rules Matrix
    # =============================================
    ws1 = wb.active
    ws1.title = "Rules Matrix"

    # Headers matching the markdown matrix
    headers = [
        "Rule", "DEP", "MOV", "CHG", "SHT", "MEL", "MOR", "RND",
        "ATK", "HIT", "DEF", "WND", "ALC", "RGN",
        "Grants", "Condition", "Value", "Target", "IMP"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Rule data structured exactly like the matrix
    # Format: [Rule, DEP, MOV, CHG, SHT, MEL, MOR, RND, ATK, HIT, DEF, WND, ALC, RGN, Grants, Condition, Value, Target, IMP, Category]
    rules_data = [
        # WEAPON RULES
        ["WEAPON RULES", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "category"],
        ["AP(X)", "", "", "", "Y", "Y", "", "", "", "", "Y", "", "", "", "Defense penalty", "Always on weapon attack", "X", "Enemy", "✅"],
        ["Blast(X)", "", "", "", "Y", "Y", "", "", "", "", "", "Y", "", "", "Multiply hits", "After hit roll", "X (capped)", "Enemy", "✅"],
        ["Deadly(X)", "", "", "", "Y", "Y", "", "", "", "", "", "Y", "", "", "Multiply wounds", "After wound calc", "X", "Enemy", "✅"],
        ["Lance", "", "", "Y", "", "Y", "", "", "", "", "Y", "", "", "", "+2 AP", "When charging", "+2", "Weapon", "✅"],
        ["Poison", "", "", "", "Y", "Y", "", "", "", "", "Y", "", "", "", "Reroll defense 6s", "On defense roll", "-", "Enemy", "✅"],
        ["Precise", "", "", "", "Y", "Y", "", "", "", "Y", "", "", "", "", "+1 to hit", "Always", "+1", "Weapon", "✅"],
        ["Reliable", "", "", "", "Y", "Y", "", "", "", "Y", "", "", "", "", "Quality becomes 2+", "Always", "2+", "Weapon", "✅"],
        ["Rending", "", "", "", "Y", "Y", "", "", "", "Y", "Y", "", "", "", "AP(4)", "On unmodified 6 to hit", "AP4", "Enemy", "⚠️"],
        ["Bane", "", "", "", "Y", "Y", "", "", "", "", "", "", "", "Y", "Bypass regeneration", "On wound", "-", "Enemy", "✅"],
        ["Impact(X)", "", "", "Y", "", "Y", "", "", "", "Y", "", "", "", "", "X extra attacks", "On charge only", "X", "Self", "⚠️"],
        ["Indirect", "", "", "", "Y", "", "", "", "", "", "Y", "", "", "", "Ignore cover", "Always", "-1 cover", "Enemy", "⚠️"],
        ["Sniper", "", "", "", "Y", "", "", "", "Y", "", "", "", "", "", "Pick target model", "When declaring", "-", "Enemy", "❌"],
        ["Lock-On", "", "", "", "Y", "Y", "", "", "", "Y", "", "", "", "", "+1 to hit vs vehicles", "vs Vehicle keyword", "+1", "Self", "❌"],
        ["Purge", "", "", "", "Y", "Y", "", "", "", "Y", "", "", "", "", "+1 to hit vs Tough(3+)", "vs Tough(3+)", "+1", "Self", "❌"],
        ["Limited", "", "", "", "Y", "Y", "", "", "Y", "", "", "", "", "", "Use once per game", "First use only", "1", "Weapon", "❌"],
        ["Linked", "", "", "", "Y", "Y", "", "", "", "Y", "", "", "", "", "Only usable with another weapon", "With paired weapon", "-", "Weapon", "❌"],

        # ATTACK MODIFIER RULES
        ["ATTACK MODIFIER RULES", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "category"],
        ["Furious", "", "", "Y", "", "Y", "", "", "", "", "", "Y", "", "", "+Sixes as extra hits", "When charging", "+6s", "Self", "✅"],
        ["PredatorFighter", "", "", "", "", "Y", "", "", "", "Y", "", "", "", "", "6s generate more attacks", "On melee 6s", "Recursive", "Self", "✅"],
        ["Relentless", "", "", "", "Y", "", "", "", "", "Y", "", "", "", "", "6s = extra hits", "Shooting >9\"", "+6s", "Self", "⚠️"],
        ["GoodShot", "", "", "", "Y", "", "", "", "", "Y", "", "", "", "", "+1 to hit", "Shooting only", "+1", "Self", "✅"],
        ["BadShot", "", "", "", "Y", "", "", "", "", "Y", "", "", "", "", "-1 to hit", "Shooting only", "-1", "Self", "✅"],
        ["Surge", "", "", "", "Y", "Y", "", "", "", "Y", "", "", "", "", "6s to hit = +1 hit", "On 6 to hit", "+1", "Self", "❌"],
        ["Thrust", "", "", "Y", "", "Y", "", "", "", "Y", "Y", "", "", "", "+1 hit, +1 AP", "When charging", "+1/+1", "Self", "❌"],
        ["VersatileAttack", "", "", "", "Y", "Y", "", "", "Y", "Y", "Y", "", "", "", "Choose: AP+1 or +1 hit", "Before attack", "+1", "Self", "✅"],
        ["PointBlankSurge", "", "", "", "Y", "", "", "", "", "Y", "", "", "", "", "6s at 0-9\" = extra hit", "Shooting 0-9\"", "+1", "Self", "❌"],

        # DEFENSE MODIFIER RULES
        ["DEFENSE MODIFIER RULES", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "category"],
        ["Tough(X)", "", "", "", "Y", "Y", "", "", "", "", "", "", "Y", "", "X wounds to kill", "Model property", "X", "Self", "✅"],
        ["Regeneration", "", "", "", "Y", "Y", "", "", "", "", "", "", "", "Y", "5+ ignore wound", "On each wound", "5+", "Self", "✅"],
        ["Shielded", "", "", "", "Y", "Y", "", "", "", "", "Y", "", "", "", "+1 Defense", "vs non-spell", "+1", "Self", "✅"],
        ["ShieldWall", "", "", "", "", "Y", "", "", "", "", "Y", "", "", "", "+1 Defense", "In melee only", "+1", "Self", "✅"],
        ["Stealth", "", "", "", "Y", "", "", "", "", "", "Y", "", "", "", "-1 to be hit", "From >12\" away", "-1 enemy", "Self", "❌"],
        ["MeleeEvasion", "", "", "", "", "Y", "", "", "", "", "Y", "", "", "", "+1 Defense", "In melee only", "+1", "Self", "✅"],
        ["MeleeShrouding", "", "", "", "", "Y", "", "", "", "", "Y", "", "", "", "+1 Defense", "In melee only", "+1", "Self", "✅"],
        ["RangedShrouding", "", "", "", "Y", "", "", "", "", "", "Y", "", "", "", "+1 Defense", "When shot", "+1", "Self", "✅"],
        ["Resistance", "", "", "", "Y", "Y", "", "", "", "", "", "", "", "Y", "6+ ignore wound (2+ vs spell)", "After wound", "6+/2+", "Self", "✅"],
        ["Protected", "", "", "", "Y", "Y", "", "", "", "", "Y", "", "", "", "6+ reduce AP by 1", "Before defense roll", "6+", "Self", "❌"],

        # WOUND MODIFIER RULES
        ["WOUND MODIFIER RULES", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "category"],
        ["Rupture", "", "", "", "Y", "Y", "", "", "", "", "", "Y", "", "Y", "+1 wound per 6, ignore regen", "On 6s to hit", "+1", "Enemy", "✅"],
        ["Shred", "", "", "", "Y", "Y", "", "", "", "", "", "Y", "", "", "+1 wound per 1 to defend", "On 1s to defend", "+1", "Enemy", "✅"],
        ["Smash", "", "", "", "Y", "Y", "", "", "", "", "", "Y", "", "Y", "Ignore regen, +Blast(3) vs Def 5+/6+", "vs Def 5+/6+", "Blast(3)", "Enemy", "✅"],

        # WOUND ALLOCATION RULES
        ["WOUND ALLOCATION RULES", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "category"],
        ["Hero", "", "", "", "Y", "Y", "", "", "", "", "", "", "Y", "", "Takes wounds last", "Allocation order", "Last", "Self", "✅"],
        ["Takedown", "", "", "", "Y", "Y", "", "", "Y", "", "", "", "Y", "", "Target single model", "Declaration", "1 model", "Enemy", "❌"],

        # REGENERATION BYPASS RULES
        ["REGENERATION BYPASS RULES", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "category"],
        ["Bane (regen bypass)", "", "", "", "Y", "Y", "", "", "", "", "", "", "", "Y", "No regeneration", "On wounds dealt", "-", "Enemy", "✅"],
        ["Rupture (regen bypass)", "", "", "", "Y", "Y", "", "", "", "", "", "Y", "", "Y", "No regeneration", "On wounds dealt", "-", "Enemy", "✅"],
        ["Unstoppable", "", "", "", "Y", "Y", "", "", "", "", "", "", "", "Y", "Ignores enemy regen", "On wounds dealt", "-", "Enemy", "⚠️"],
        ["BaneInMelee", "", "", "", "", "Y", "", "", "", "", "", "", "", "Y", "All melee attacks have Bane", "Melee only", "-", "Self", "✅"],

        # MORALE RULES
        ["MORALE RULES", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "category"],
        ["Fearless", "", "", "", "", "", "Y", "", "", "", "", "", "", "", "Reroll failed morale", "On morale failure", "Reroll", "Self", "✅"],
        ["MoraleBoost", "", "", "", "", "", "Y", "", "", "", "", "", "", "", "+1 to morale tests", "Always", "+1", "Self", "✅"],
        ["NoRetreat", "", "", "", "", "", "Y", "", "", "", "", "", "", "", "Take wounds instead of Shaken", "On morale failure", "wounds", "Self", "✅"],
        ["Battleborn", "", "", "", "", "", "", "Y", "", "", "", "", "", "", "4+ to rally from Shaken", "Round start", "4+", "Self", "✅"],
        ["Fear(X)", "", "", "", "", "Y", "Y", "", "", "", "", "", "", "", "Count as +X wounds", "In melee morale", "+X", "Self", "❌"],
        ["HoldTheLine", "", "", "", "", "", "Y", "", "", "", "", "", "", "", "Reroll failed morale", "On morale failure", "Reroll", "Self", "✅"],

        # COMBAT ORDER RULES
        ["COMBAT ORDER RULES", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "category"],
        ["Counter", "", "", "", "", "Y", "", "", "Y", "", "", "", "", "", "Strike first when charged", "When receiving charge", "First", "Self", "❌"],
        ["HitAndRun", "", "", "", "", "Y", "", "Y", "", "", "", "", "", "", "Retreat after combat", "End of melee", "-", "Self", "❌"],
        ["SelfDestruct", "", "", "", "", "Y", "", "", "", "", "", "", "Y", "", "Deal X hits when killed", "On model death", "X", "Enemy", "❌"],

        # MOVEMENT/DEPLOYMENT RULES
        ["MOVEMENT/DEPLOYMENT RULES", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "category"],
        ["Scout", "Y", "", "", "", "", "", "", "", "", "", "", "", "", "Deploy 12\" ahead", "Deployment", "+12\"", "Self", "❌"],
        ["Ambush", "Y", "", "", "", "", "", "", "", "", "", "", "", "", "Deploy >9\" from enemy", "Deployment", "Reserve", "Self", "❌"],
        ["Fast", "", "Y", "Y", "", "", "", "", "", "", "", "", "", "", "9\" move", "Movement", "9\"", "Self", "❌"],
        ["Slow", "", "Y", "Y", "", "", "", "", "", "", "", "", "", "", "4\" move", "Movement", "4\"", "Self", "❌"],
        ["Agile", "", "Y", "Y", "", "", "", "", "", "", "", "", "", "", "+1\" advance, +2\" rush/charge", "Movement", "+1/+2", "Self", "❌"],
        ["Flying", "", "Y", "Y", "", "", "", "", "", "", "", "", "", "", "Ignore terrain/units", "Movement", "-", "Self", "❌"],
        ["Strider", "", "Y", "Y", "", "", "", "", "", "", "", "", "", "", "Ignore difficult terrain", "Movement", "-", "Self", "❌"],
        ["RapidCharge", "", "", "Y", "", "", "", "", "", "", "", "", "", "", "+4\" charge range", "Charge only", "+4\"", "Self", "❌"],

        # MAGIC/SPECIAL RULES
        ["MAGIC/SPECIAL RULES", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "category"],
        ["Casting(X)", "", "", "", "", "", "", "Y", "", "", "", "", "", "", "Cast X spells per round", "Casting phase", "X", "Self", "❌"],
        ["Devout", "", "", "", "", "", "", "", "", "", "", "", "", "", "Faction rule", "Varies", "-", "Self", "❌"],
        ["PiercingAssault", "", "", "Y", "", "Y", "", "", "", "", "Y", "", "", "", "AP(1) on melee charge", "When charging", "AP(1)", "Self", "✅"],
    ]

    row_idx = 2
    for row_data in rules_data:
        is_category = len(row_data) > 19 and row_data[19] == "category"

        if is_category:
            # Category row - merge and style
            cell = ws1.cell(row=row_idx, column=1, value=row_data[0])
            cell.font = category_font
            cell.fill = category_fill
            cell.border = thin_border
            ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=19)
            row_idx += 1
            continue

        for col_idx, value in enumerate(row_data[:19], 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col_idx <= 14 else 'left', vertical='center', wrap_text=True)

            # Phase columns (Y markers)
            if col_idx >= 2 and col_idx <= 14 and value == "Y":
                cell.fill = phase_yes_fill

            # Implementation status column
            if col_idx == 19:
                if "✅" in str(value):
                    cell.fill = implemented_fill
                elif "⚠️" in str(value):
                    cell.fill = partial_fill
                elif "❌" in str(value):
                    cell.fill = not_impl_fill

        row_idx += 1

    # Set column widths
    ws1.column_dimensions['A'].width = 18
    for col in range(2, 15):
        ws1.column_dimensions[get_column_letter(col)].width = 5
    ws1.column_dimensions['O'].width = 30
    ws1.column_dimensions['P'].width = 25
    ws1.column_dimensions['Q'].width = 12
    ws1.column_dimensions['R'].width = 10
    ws1.column_dimensions['S'].width = 6

    # Freeze first row
    ws1.freeze_panes = 'A2'

    # =============================================
    # SHEET 2: Aura Rules
    # =============================================
    ws2 = wb.create_sheet("Aura Rules")

    aura_headers = ["Aura Name", "Grants Rule", "Effect Description", "Trigger", "Target", "IMP"]
    for col, header in enumerate(aura_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    aura_data = [
        ["Furious Aura", "Furious", "Grants Furious - extra hits on 6s when charging", "Always", "Nearby units", "✅"],
        ["Shielded Aura", "Shielded", "Grants +1 defense vs non-spell hits", "Always", "Nearby units", "✅"],
        ["Regeneration Aura", "Regeneration", "Grants 5+ save to ignore wounds", "Always", "Nearby units", "✅"],
        ["Relentless Aura", "Relentless", "Grants extra hits on 6s when shooting >9\"", "Always", "Nearby units", "⚠️"],
        ["Scout Aura", "Scout", "Grants Scout deployment ability", "Deployment", "Nearby units", "❌"],
        ["Stealth Aura", "Stealth", "Grants -1 to be hit from >12\" away", "Always", "Nearby units", "❌"],
        ["Counter-Attack Aura", "Counter", "Grants strike first when charged", "When charged", "Nearby units", "❌"],
        ["Fearless Aura", "Fearless", "Grants reroll failed morale tests", "Morale phase", "Nearby units", "✅"],
        ["Ambush Aura", "Ambush", "Grants ability to deploy >9\" from enemy", "Deployment", "Nearby units", "❌"],
        ["Psychic Blast", "Deals damage", "Deals 3 wounds with AP(1)", "When attacking", "Enemy unit", "✅"],
    ]

    for row_idx, row_data in enumerate(aura_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col_idx == 6:
                if "✅" in value:
                    cell.fill = implemented_fill
                elif "⚠️" in value:
                    cell.fill = partial_fill
                elif "❌" in value:
                    cell.fill = not_impl_fill

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 45
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 8

    # =============================================
    # SHEET 3: Faction Army Rules
    # =============================================
    ws3 = wb.create_sheet("Faction Army Rules")

    faction_headers = ["Faction", "Rule Name", "Rule Type", "Effect Description", "Trigger Timing"]
    for col, header in enumerate(faction_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    faction_data = [
        ["Alien Hives", "Hive Bond (MoraleBoost)", "ArmyWide", "+1 to morale tests for all units", "Always"],
        ["Alien Hives", "Psychic Blast", "Aura", "Deals 3 wounds with AP(1)", "WhenAttacking"],
        ["Battle Brothers", "Battleborn", "ArmyWide", "4+ to stop being Shaken at round start", "StartOfRound"],
        ["Battle Brothers", "Furious Aura", "Aura", "Grants Furious to nearby units", "Always"],
        ["Blessed Sisters", "Devout", "ArmyWide", "Faction-specific bonus", "Always"],
        ["Blessed Sisters", "Furious Aura", "Aura", "Grants Furious to nearby units", "Always"],
        ["Blood Brothers", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["Blood Brothers", "Furious", "Special", "Extra hits on 6s when charging", "OnCharge"],
        ["Blood Prime Brothers", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["Blood Prime Brothers", "Agile", "Special", "+1\" advance, +2\" rush/charge", "Always"],
        ["Change Disciples", "Resistance", "ArmyWide", "6+ ignore wounds (2+ vs spells)", "WhenDefending"],
        ["Custodian Brothers", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["Custodian Brothers", "Shielded", "Special", "+1 defense vs non-spell hits", "WhenDefending"],
        ["DAO Union", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["DAO Union", "Good Shot", "Special", "+1 to hit when shooting", "WhenAttacking"],
        ["Dark Brothers", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["Dark Brothers", "Stealth", "Special", "-1 to be hit from >12\"", "WhenDefending"],
        ["Dark Elf Raiders", "Agile", "ArmyWide", "+1\" advance, +2\" rush/charge", "Always"],
        ["Dark Elf Raiders", "Hit and Run", "Special", "Can retreat after fighting", "EndOfMelee"],
        ["Dark Prime Brothers", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["Dwarf Guilds", "Hold The Line", "ArmyWide", "Reroll failed morale tests", "Always"],
        ["Dwarf Guilds", "Shielded", "Special", "+1 defense vs non-spell hits", "WhenDefending"],
        ["Elven Jesters", "Agile", "ArmyWide", "+1\" advance, +2\" rush/charge", "Always"],
        ["Eternal Dynasty", "No Retreat", "ArmyWide", "Can't be shaken/routed, take wounds", "Always"],
        ["Eternal Dynasty", "Regeneration", "Special", "5+ save to ignore wound", "WhenDefending"],
        ["Goblin Reclaimers", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["Havoc Brothers", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["Havoc Brothers", "Furious", "Special", "Extra hits on 6s when charging", "OnCharge"],
        ["High Elf Fleets", "Agile", "ArmyWide", "+1\" advance, +2\" rush/charge", "Always"],
        ["High Elf Fleets", "Counter", "Special", "Strikes first when charged", "OnBeingCharged"],
        ["Human Defense Force", "Hold The Line", "ArmyWide", "Reroll failed morale tests", "Always"],
        ["Human Inquisition", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["Human Inquisition", "Devout", "Special", "Faction-specific bonus", "Always"],
        ["Infected Colonies", "Fearless", "ArmyWide", "Reroll failed morale tests", "Always"],
        ["Infected Colonies", "Regeneration", "Special", "5+ save to ignore wound", "WhenDefending"],
        ["Jackals", "Agile", "ArmyWide", "+1\" advance, +2\" rush/charge", "Always"],
        ["Jackals", "Hit and Run", "Special", "Can retreat after fighting", "EndOfMelee"],
        ["Knight Brothers", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["Knight Prime Brothers", "Battleborn", "ArmyWide", "4+ to stop being Shaken", "StartOfRound"],
        ["Lust Disciples", "Agile", "ArmyWide", "+1\" advance, +2\" rush/charge", "Always"],
    ]

    for row_idx, row_data in enumerate(faction_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 45
    ws3.column_dimensions['E'].width = 18

    # =============================================
    # SHEET 4: Implementation Status Summary
    # =============================================
    ws4 = wb.create_sheet("Implementation Status")

    # Count rules (excluding category rows)
    impl_rules = [r for r in rules_data if len(r) <= 19 or r[19] != "category"]
    implemented = [r[0] for r in impl_rules if len(r) > 18 and "✅" in str(r[18])]
    partial = [r[0] for r in impl_rules if len(r) > 18 and "⚠️" in str(r[18])]
    not_impl = [r[0] for r in impl_rules if len(r) > 18 and "❌" in str(r[18])]
    total = len(implemented) + len(partial) + len(not_impl)

    status_headers = ["Status", "Count", "Percentage", "Rules"]
    for col, header in enumerate(status_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    status_data = [
        ["✅ Fully Implemented", str(len(implemented)), f"{len(implemented)/total*100:.1f}%", ", ".join(implemented)],
        ["⚠️ Partially Implemented", str(len(partial)), f"{len(partial)/total*100:.1f}%", ", ".join(partial)],
        ["❌ Not Implemented", str(len(not_impl)), f"{len(not_impl)/total*100:.1f}%", ", ".join(not_impl)],
        ["TOTAL", str(total), "100%", ""],
    ]

    for row_idx, row_data in enumerate(status_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col_idx == 1:
                if "✅" in value:
                    cell.fill = implemented_fill
                elif "⚠️" in value:
                    cell.fill = partial_fill
                elif "❌" in value:
                    cell.fill = not_impl_fill

    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 10
    ws4.column_dimensions['C'].width = 12
    ws4.column_dimensions['D'].width = 100

    # =============================================
    # SHEET 5: Combat Resolution Flow
    # =============================================
    ws5 = wb.create_sheet("Combat Resolution Flow")

    ws5.cell(row=1, column=1, value="COMBAT RESOLUTION FLOW").font = Font(bold=True, size=14)
    ws5.merge_cells('A1:C1')

    flow_data = [
        ["", "", ""],
        ["PHASE", "STEP", "RULES APPLIED"],
        ["ROUND START", "", ""],
        ["", "Battleborn check", "Battleborn (4+ to rally)"],
        ["", "", ""],
        ["ATTACK DECLARATION", "", ""],
        ["", "VersatileAttack choice", "VersatileAttack (AP+1 or +1 hit)"],
        ["", "Target selection", "Sniper, Takedown"],
        ["", "Initiative swap", "Counter (swap initiative)"],
        ["", "", ""],
        ["HIT ROLL", "Quality Test", ""],
        ["", "Base modifiers", "Reliable (2+), Precise (+1)"],
        ["", "Shooting modifiers", "GoodShot (+1), BadShot (-1)"],
        ["", "Conditional modifiers", "Lock-On (+1 vs vehicles), Purge (+1 vs Tough 3+)"],
        ["", "Shaken penalty", "-1 to hit"],
        ["", "Track 6s rolled", "For Furious, Rending, Rupture, etc."],
        ["", "", ""],
        ["HIT MODIFIERS", "", ""],
        ["", "Charge bonuses", "Furious (+6s), Impact(X) (+X attacks)"],
        ["", "Blast multiplier", "Blast(X) (multiply by min(X, models))"],
        ["", "Shooting bonuses", "Relentless (+6s >9\"), Surge (+1 per 6)"],
        ["", "Recursive attacks", "PredatorFighter (6s generate more)"],
        ["", "", ""],
        ["CALCULATE AP", "", ""],
        ["", "Base AP", "From weapon"],
        ["", "Charge bonuses", "Lance (+2), Thrust (+1), PiercingAssault (+1)"],
        ["", "Hit-based", "Rending (AP4 on 6s)"],
        ["", "Choice-based", "VersatileAttack (+1 if chosen)"],
        ["", "Defender reduction", "Protected (6+ to reduce AP by 1)"],
        ["", "", ""],
        ["DEFENSE ROLL", "Save Test", ""],
        ["", "Base Defense", "From model"],
        ["", "Cover bonus", "+1"],
        ["", "Defense bonuses", "Shielded (+1), ShieldWall (+1 melee)"],
        ["", "Evasion bonuses", "MeleeEvasion (+1), MeleeShrouding (+1)"],
        ["", "Ranged defense", "RangedShrouding (+1), Stealth (+1 >12\")"],
        ["", "Ignore cover", "Indirect"],
        ["", "Reroll defense", "Poison (reroll 6s)"],
        ["", "Track 1s rolled", "For Shred"],
        ["", "", ""],
        ["WOUND CALCULATION", "", ""],
        ["", "Base wounds", "Hits - Saves"],
        ["", "Multipliers", "Deadly(X)"],
        ["", "Bonus wounds", "Rupture (+1 per 6), Shred (+1 per 1)"],
        ["", "Smash bonus", "+Blast(3) vs Def 5+/6+"],
        ["", "", ""],
        ["RESISTANCE", "", ""],
        ["", "Resistance roll", "6+ to ignore (2+ vs spells)"],
        ["", "", ""],
        ["WOUND ALLOCATION", "", ""],
        ["", "Order", "Non-tough → Tough (most wounded) → Heroes"],
        ["", "Check thresholds", "Tough(X)"],
        ["", "", ""],
        ["REGENERATION", "", ""],
        ["", "Regeneration roll", "5+ to ignore each wound"],
        ["", "Bypass rules", "Bane, Rupture, Unstoppable, BaneInMelee"],
        ["", "", ""],
        ["MODEL DEATH", "", ""],
        ["", "Death triggers", "SelfDestruct (deal X hits)"],
        ["", "", ""],
        ["MORALE CHECK", "", ""],
        ["", "Base Morale", "From unit"],
        ["", "Bonuses", "MoraleBoost (+1)"],
        ["", "Fear modifier", "Fear(X) (count as +X wounds)"],
        ["", "Reroll failures", "Fearless, HoldTheLine"],
        ["", "Alternative", "NoRetreat (wounds instead of Shaken)"],
        ["", "", ""],
        ["ROUND END", "", ""],
        ["", "Post-combat", "HitAndRun (retreat option)"],
    ]

    for row_idx, row_data in enumerate(flow_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col_idx == 1 and value and not value.startswith(" "):
                cell.font = Font(bold=True)
                cell.fill = category_fill

    ws5.column_dimensions['A'].width = 22
    ws5.column_dimensions['B'].width = 25
    ws5.column_dimensions['C'].width = 50

    # =============================================
    # SHEET 6: Legend
    # =============================================
    ws6 = wb.create_sheet("Legend")

    ws6.cell(row=1, column=1, value="LEGEND").font = Font(bold=True, size=14)

    legend_data = [
        ["", ""],
        ["PHASE COLUMNS", ""],
        ["DEP", "Deployment phase"],
        ["MOV", "Movement phase"],
        ["CHG", "Charge declaration/movement"],
        ["SHT", "Shooting phase"],
        ["MEL", "Melee phase"],
        ["MOR", "Morale check"],
        ["RND", "Round start/end"],
        ["", ""],
        ["SUB-STEP COLUMNS", ""],
        ["ATK", "Attack declaration (choosing targets)"],
        ["HIT", "Hit roll (Quality test)"],
        ["DEF", "Defense roll (Save test)"],
        ["WND", "Wound calculation/modification"],
        ["ALC", "Wound allocation to models"],
        ["RGN", "Regeneration rolls"],
        ["", ""],
        ["EFFECT COLUMNS", ""],
        ["Grants", "What the rule provides (modifier, extra dice, etc.)"],
        ["Condition", "When/how the effect triggers"],
        ["Value", "Numeric value if applicable"],
        ["Target", "Who is affected (Self, Enemy, Weapon)"],
        ["", ""],
        ["IMPLEMENTATION STATUS", ""],
        ["✅", "Fully implemented in codebase"],
        ["⚠️", "Partially implemented"],
        ["❌", "Not implemented"],
    ]

    for row_idx, row_data in enumerate(legend_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws6.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='center')
            if col_idx == 1 and value and value.isupper():
                cell.font = Font(bold=True)
            if value == "✅":
                cell.fill = implemented_fill
            elif value == "⚠️":
                cell.fill = partial_fill
            elif value == "❌":
                cell.fill = not_impl_fill

    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 50

    # Save workbook
    output_path = "/home/user/Science-Battle-Simulator/chunk_sim_rules_review.xlsx"
    wb.save(output_path)
    print(f"Excel file created: {output_path}")
    print(f"Total rules: {total}")
    print(f"  Implemented: {len(implemented)}")
    print(f"  Partial: {len(partial)}")
    print(f"  Not implemented: {len(not_impl)}")
    return output_path

if __name__ == "__main__":
    create_rules_sheet()
