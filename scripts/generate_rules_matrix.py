#!/usr/bin/env python3
"""Generate Excel spreadsheet for special rules engine configuration.

Reads faction-specific rules from Faction Specific Army Rules.xlsx and
generates an engine-ready configuration spreadsheet.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl
from collections import defaultdict
import re

def extract_rules_from_faction_file():
    """Extract all unique rules from the faction rules Excel file."""
    wb = openpyxl.load_workbook('/home/user/Science-Battle-Simulator/Faction Specific Army Rules.xlsx')
    ws = wb['Sheet1']

    rules = {}
    armies_per_rule = defaultdict(set)

    for row in ws.iter_rows(min_row=2, values_only=True):
        army, rule_type, rule_name, rule_text, type_col, once_per_game, once_per_activation = row
        if rule_name and rule_text:
            if rule_name not in rules:
                rules[rule_name] = {
                    'text': str(rule_text),
                    'type': type_col or rule_type or '',
                    'once_per_game': bool(once_per_game),
                    'once_per_activation': bool(once_per_activation)
                }
            armies_per_rule[rule_name].add(army)

    return rules, armies_per_rule

def parse_rule_for_engine(name, info):
    """Parse a rule's text to determine engine configuration."""
    text = info['text'].lower()
    rule_type = (info['type'] or '').lower()

    # Default values
    phases = {'DEP': '', 'MOV': '', 'CHG': '', 'SHT': '', 'MEL': '', 'MOR': '', 'RND': ''}
    trigger_step = ''
    trigger_condition = 'always'
    effect_type = ''
    effect_target = ''
    effect_value = ''
    target_unit = 'SELF'
    priority = '10'

    # Determine phases based on keywords in text
    if any(x in text for x in ['deploy', 'ambush', 'scout', 'infiltrate']):
        phases['DEP'] = 'Y'
    if any(x in text for x in ['move', 'advance', 'rush', 'fast', 'slow', 'strider', 'flying']):
        phases['MOV'] = 'Y'
    if any(x in text for x in ['charge', 'charging', 'impact']):
        phases['CHG'] = 'Y'
    if any(x in text for x in ['shoot', 'shooting', 'range', 'ranged']):
        phases['SHT'] = 'Y'
    if any(x in text for x in ['melee', 'close combat', 'fight']):
        phases['MEL'] = 'Y'
    if any(x in text for x in ['morale', 'shaken', 'routed', 'fearless', 'fear']):
        phases['MOR'] = 'Y'
    if any(x in text for x in ['beginning of', 'end of', 'round', 'activation']):
        phases['RND'] = 'Y'

    # If nothing detected, check type column
    if not any(phases.values()):
        if 'weapon' in rule_type:
            phases['SHT'] = 'Y'
            phases['MEL'] = 'Y'
        elif 'defense' in rule_type:
            phases['SHT'] = 'Y'
            phases['MEL'] = 'Y'
        elif 'movement' in rule_type:
            phases['MOV'] = 'Y'
        elif 'aura' in rule_type:
            phases['RND'] = 'Y'

    # Determine trigger step and effects based on keywords
    # Hit modifiers
    if '+1 to hit' in text or 'gets +1 to hit' in text:
        trigger_step = 'HIT_ROLL'
        effect_type = 'ADD'
        effect_target = 'quality_mod'
        effect_value = '1'
    elif '-1 to hit' in text or 'gets -1 to hit' in text:
        trigger_step = 'HIT_ROLL'
        effect_type = 'ADD'
        effect_target = 'quality_mod'
        effect_value = '-1'
    elif 'unmodified roll of 6 to hit' in text or 'unmodified results of 6 to hit' in text:
        trigger_step = 'AFTER_HIT'
        trigger_condition = 'unmodified_6_hit'
        if 'extra hit' in text or '+1 attack' in text or 'may roll +1 attack' in text:
            effect_type = 'ADD'
            effect_target = 'hits'
            effect_value = 'sixes_rolled'
        elif 'ap (+' in text or 'ap(+' in text:
            match = re.search(r'ap\s*\(\+?(\d+)\)', text)
            if match:
                effect_type = 'ADD'
                effect_target = 'ap'
                effect_value = match.group(1)
        elif 'extra wound' in text:
            effect_type = 'ADD'
            effect_target = 'wounds'
            effect_value = 'sixes_rolled'

    # Defense modifiers
    elif '+1 to defense' in text or 'gets +1 to defense' in text or '+1 defense' in text:
        trigger_step = 'DEF_ROLL'
        effect_type = 'ADD'
        effect_target = 'defense_mod'
        effect_value = '1'
    elif '-1 to defense' in text or 'gets -1 to defense' in text or '-1 defense' in text:
        trigger_step = 'DEF_ROLL'
        effect_type = 'ADD'
        effect_target = 'defense_mod'
        effect_value = '-1'
        target_unit = 'ENEMY'

    # Morale modifiers
    elif '+1 to morale' in text:
        trigger_step = 'MORALE'
        effect_type = 'ADD'
        effect_target = 'morale_mod'
        effect_value = '1'
    elif 'fearless' in text and ('reroll' in text or 'get fearless' in text):
        trigger_step = 'MORALE'
        effect_type = 'REROLL'
        effect_target = 'morale_test'
        effect_value = 'true'

    # Movement modifiers
    elif 'moves +' in text or '+1"' in text or '+2"' in text:
        trigger_step = 'MOVEMENT'
        effect_type = 'ADD'
        effect_target = 'move_distance'
        match = re.search(r'\+(\d+)"', text)
        if match:
            effect_value = match.group(1)

    # AP modifiers
    elif 'ap (+' in text or 'ap(+' in text:
        trigger_step = 'CALC_AP'
        effect_type = 'ADD'
        effect_target = 'ap'
        match = re.search(r'ap\s*\(\+?(\d+)\)', text)
        if match:
            effect_value = match.group(1)
    elif 'ap (' in text or 'ap(' in text:
        trigger_step = 'CALC_AP'
        effect_type = 'SET'
        effect_target = 'ap'
        match = re.search(r'ap\s*\((\d+)\)', text)
        if match:
            effect_value = match.group(1)

    # Blast/hits
    elif 'blast' in text:
        trigger_step = 'AFTER_HIT'
        effect_type = 'MULTIPLY'
        effect_target = 'hits'
        match = re.search(r'blast\s*\((\d+)\)', text)
        if match:
            effect_value = f"min({match.group(1)}, enemy_models)"

    # Deadly
    elif 'deadly' in text:
        trigger_step = 'AFTER_WOUND'
        effect_type = 'MULTIPLY'
        effect_target = 'wounds'
        match = re.search(r'deadly\s*\((\d+)\)', text)
        if match:
            effect_value = match.group(1)

    # Regeneration/ignore wounds
    elif 'ignore' in text and 'wound' in text:
        if 'regeneration' in text:
            trigger_step = 'REGEN'
            effect_type = 'NEGATE'
            effect_target = 'regeneration'
            effect_value = 'true'
            target_unit = 'ENEMY'
        else:
            trigger_step = 'REGEN'
            effect_type = 'ROLL'
            match = re.search(r'(\d)\+', text)
            if match:
                effect_target = 'ignore_wound'
                effect_value = f"{match.group(1)}+"

    # Aura effects
    elif 'pick one' in text or 'pick up to' in text:
        trigger_step = 'PRE_ATTACK'
        effect_type = 'GRANT'
        if 'enemy' in text:
            target_unit = 'ENEMY'
        if 'friendly' in text:
            target_unit = 'FRIENDLY'
        # Try to find what is granted
        if 'takes' in text and 'hit' in text:
            effect_type = 'DAMAGE'
            match = re.search(r'takes?\s+(\d+)\s+hits?', text)
            if match:
                effect_value = match.group(1)
                effect_target = 'direct_hits'

    # Takes X hits (direct damage)
    elif 'takes' in text and 'hit' in text:
        trigger_step = 'DIRECT_DAMAGE'
        effect_type = 'DAMAGE'
        target_unit = 'ENEMY'
        match = re.search(r'takes?\s+(\d+)\s+hits?', text)
        if match:
            effect_target = 'direct_hits'
            effect_value = match.group(1)

    # Counter
    elif 'counter' in text and 'strike' in text:
        trigger_step = 'COMBAT_ORDER'
        trigger_condition = 'is_charged'
        effect_type = 'SET'
        effect_target = 'strike_order'
        effect_value = 'first'

    # Ambush
    elif 'ambush' in text:
        phases['DEP'] = 'Y'
        trigger_step = 'DEPLOYMENT'
        effect_type = 'SET'
        effect_target = 'deploy_type'
        effect_value = 'reserve'

    # Scout/Infiltrate
    elif 'scout' in text or 'infiltrate' in text:
        phases['DEP'] = 'Y'
        trigger_step = 'DEPLOYMENT'
        effect_type = 'ADD'
        effect_target = 'deploy_distance'
        effect_value = '12'

    # Flying
    elif 'flying' in text or 'fly over' in text:
        phases['MOV'] = 'Y'
        trigger_step = 'MOVEMENT'
        effect_type = 'NEGATE'
        effect_target = 'terrain_blocking'
        effect_value = 'true'

    # Strider
    elif 'strider' in text or 'ignore' in text and 'terrain' in text:
        phases['MOV'] = 'Y'
        trigger_step = 'MOVEMENT'
        effect_type = 'NEGATE'
        effect_target = 'difficult_terrain'
        effect_value = 'true'

    # Cover related
    elif 'cover' in text and 'ignore' in text:
        trigger_step = 'DEF_ROLL'
        effect_type = 'NEGATE'
        effect_target = 'cover_bonus'
        effect_value = 'true'
        target_unit = 'ENEMY'

    # Shielded
    elif 'shielded' in text:
        trigger_step = 'DEF_ROLL'
        trigger_condition = 'attack_not_spell'
        effect_type = 'ADD'
        effect_target = 'defense_mod'
        effect_value = '1'

    # Tough
    elif 'tough' in text:
        trigger_step = 'ALLOCATE'
        effect_type = 'SET'
        effect_target = 'wounds_to_kill'
        match = re.search(r'tough\s*\((\d+)\)', text)
        if match:
            effect_value = match.group(1)

    # Furious
    elif 'furious' in text:
        trigger_step = 'AFTER_HIT'
        trigger_condition = 'is_charging'
        effect_type = 'ADD'
        effect_target = 'hits'
        effect_value = 'sixes_rolled'

    # Relentless
    elif 'relentless' in text:
        trigger_step = 'AFTER_HIT'
        trigger_condition = 'range_over_9'
        effect_type = 'ADD'
        effect_target = 'hits'
        effect_value = 'sixes_rolled'

    # Handle once per game/activation
    if info['once_per_game']:
        trigger_condition = f"once_per_game AND {trigger_condition}" if trigger_condition != 'always' else 'once_per_game'
    if info['once_per_activation']:
        trigger_condition = f"once_per_activation AND {trigger_condition}" if trigger_condition != 'always' else 'once_per_activation'

    # If we still don't have phases, default to combat
    if not any(phases.values()):
        phases['SHT'] = 'Y'
        phases['MEL'] = 'Y'

    return {
        'phases': phases,
        'trigger_step': trigger_step,
        'trigger_condition': trigger_condition,
        'effect_type': effect_type,
        'effect_target': effect_target,
        'effect_value': effect_value,
        'target_unit': target_unit,
        'priority': priority
    }

def create_rule_id(name):
    """Create a code-friendly rule ID from the rule name."""
    # Remove parentheses content
    clean = re.sub(r'\s*\([^)]*\)', '', name)
    # Replace spaces and special chars with underscores
    clean = re.sub(r'[^a-zA-Z0-9]', '_', clean)
    # Remove consecutive underscores
    clean = re.sub(r'_+', '_', clean)
    # Remove leading/trailing underscores
    clean = clean.strip('_')
    return clean.upper()

def get_base_rules():
    """Return the base special rules from the core game system."""
    # These are the base rules defined in types.hpp RuleId enum
    # Format: (rule_id, name, description, DEP, MOV, CHG, SHT, MEL, MOR, RND,
    #          trigger_step, trigger_condition, effect_type, effect_target, effect_value, target_unit, priority, notes)
    return [
        # Weapon rules (affect attacks)
        ("AP", "AP(X)", "Armor Piercing - adds X to the target's defense roll requirement",
         "", "", "", "Y", "Y", "", "", "CALC_AP", "always", "ADD", "defense_target", "X", "ENEMY", "10", "Base Rule - Weapon"),
        ("BLAST", "Blast(X)", "Each successful hit is multiplied, up to the number of models in the target unit",
         "", "", "", "Y", "Y", "", "", "AFTER_HIT", "always", "MULTIPLY", "hits", "min(X, enemy_models)", "ENEMY", "20", "Base Rule - Weapon"),
        ("DEADLY", "Deadly(X)", "Each wound caused is multiplied by X",
         "", "", "", "Y", "Y", "", "", "AFTER_WOUND", "always", "MULTIPLY", "wounds", "X", "ENEMY", "10", "Base Rule - Weapon"),
        ("LANCE", "Lance", "Gets +2 AP when charging",
         "", "", "Y", "", "Y", "", "", "CALC_AP", "is_charging", "ADD", "ap", "2", "WEAPON", "10", "Base Rule - Weapon"),
        ("POISON", "Poison", "Defender must reroll successful defense rolls of 6",
         "", "", "", "Y", "Y", "", "", "DEF_ROLL", "always", "REROLL", "defense_6s", "true", "ENEMY", "10", "Base Rule - Weapon"),
        ("PRECISE", "Precise", "Gets +1 to hit rolls",
         "", "", "", "Y", "Y", "", "", "HIT_ROLL", "always", "ADD", "quality_mod", "1", "WEAPON", "10", "Base Rule - Weapon"),
        ("RELIABLE", "Reliable", "Attacks always hit on 2+",
         "", "", "", "Y", "Y", "", "", "HIT_ROLL", "always", "SET", "effective_quality", "2", "WEAPON", "5", "Base Rule - Weapon"),
        ("RENDING", "Rending", "Unmodified rolls of 6 to hit get AP(4)",
         "", "", "", "Y", "Y", "", "", "CALC_AP", "unmodified_6_hit", "SET", "ap", "4", "WEAPON", "20", "Base Rule - Weapon"),
        ("BANE", "Bane", "Wounds caused cannot be ignored by Regeneration",
         "", "", "", "Y", "Y", "", "", "REGEN", "always", "NEGATE", "regeneration", "true", "ENEMY", "10", "Base Rule - Weapon"),
        ("IMPACT", "Impact(X)", "Gets X extra attacks when charging",
         "", "", "Y", "", "Y", "", "", "CALC_ATTACKS", "is_charging", "ADD", "attacks", "X", "WEAPON", "10", "Base Rule - Weapon"),
        ("INDIRECT", "Indirect", "May target enemies out of line of sight, ignores cover",
         "", "", "", "Y", "", "", "", "DEF_ROLL", "always", "NEGATE", "cover_bonus", "true", "ENEMY", "10", "Base Rule - Weapon"),
        ("SNIPER", "Sniper", "May pick which model in the target unit takes the wounds",
         "", "", "", "Y", "", "", "", "ALLOCATE", "always", "REPLACE", "allocation_order", "chosen_model", "ENEMY", "10", "Base Rule - Weapon"),
        ("LOCK_ON", "Lock-On", "Gets +1 to hit against targets with the Vehicle keyword",
         "", "", "", "Y", "Y", "", "", "HIT_ROLL", "target_is_vehicle", "ADD", "quality_mod", "1", "SELF", "10", "Base Rule - Weapon"),
        ("PURGE", "Purge", "Gets +1 to hit against targets with Tough(3) or higher",
         "", "", "", "Y", "Y", "", "", "HIT_ROLL", "target_has_tough_3plus", "ADD", "quality_mod", "1", "SELF", "10", "Base Rule - Weapon"),
        ("LIMITED", "Limited", "This weapon may only be used once per game",
         "", "", "", "Y", "Y", "", "", "CALC_ATTACKS", "not_used_this_game", "SET", "weapon_available", "false_after_use", "WEAPON", "1", "Base Rule - Weapon"),

        # Defense rules
        ("REGENERATION", "Regeneration", "When taking wounds, roll one die for each. On 5+ the wound is ignored",
         "", "", "", "Y", "Y", "", "", "REGEN", "always", "ROLL", "ignore_wound", "5+", "SELF", "10", "Base Rule - Defense"),
        ("TOUGH", "Tough(X)", "This model must take X wounds before being removed as a casualty",
         "", "", "", "Y", "Y", "", "", "ALLOCATE", "always", "SET", "wounds_to_kill", "X", "SELF", "1", "Base Rule - Defense"),
        ("PROTECTED", "Protected", "Roll one die for each hit. On 6+ reduce the AP of that hit by 1",
         "", "", "", "Y", "Y", "", "", "CALC_AP", "always", "ROLL", "reduce_ap", "6+ for -1", "SELF", "5", "Base Rule - Defense"),
        ("STEALTH", "Stealth", "Enemies shooting at this unit from more than 12\" away get -1 to hit",
         "", "", "", "Y", "", "", "", "HIT_ROLL", "range_over_12", "ADD", "enemy_quality_mod", "-1", "SELF", "10", "Base Rule - Defense"),
        ("SHIELD_WALL", "Shield Wall", "Gets +1 to defense rolls in melee",
         "", "", "", "", "Y", "", "", "DEF_ROLL", "phase_is_melee", "ADD", "defense_mod", "1", "SELF", "10", "Base Rule - Defense"),
        ("SHIELDED", "Shielded", "Gets +1 to defense rolls against hits that are not from spells",
         "", "", "", "Y", "Y", "", "", "DEF_ROLL", "attack_not_spell", "ADD", "defense_mod", "1", "SELF", "10", "Base Rule - Defense"),
        ("RESISTANCE", "Resistance", "When taking wounds, roll one die for each. On 6+ it is ignored (2+ vs spells)",
         "", "", "", "Y", "Y", "", "", "REGEN", "always", "ROLL", "ignore_wound", "6+ OR 2+_vs_spell", "SELF", "20", "Base Rule - Defense"),

        # Morale rules
        ("FEARLESS", "Fearless", "May reroll failed morale tests",
         "", "", "", "", "", "Y", "", "MORALE", "morale_failed", "REROLL", "morale_test", "true", "SELF", "10", "Base Rule - Morale"),
        ("MORALE_BOOST", "Morale Boost", "Gets +1 to morale test rolls",
         "", "", "", "", "", "Y", "", "MORALE", "always", "ADD", "morale_mod", "1", "SELF", "10", "Base Rule - Morale"),
        ("NO_RETREAT", "No Retreat", "When failing morale, don't become Shaken. Instead roll dice equal to the margin of failure, taking a wound for each 1-3",
         "", "", "", "", "", "Y", "", "MORALE", "would_become_shaken", "REPLACE", "shaken_result", "roll_for_wounds", "SELF", "10", "Base Rule - Morale"),
        ("BATTLEBORN", "Battleborn", "At the beginning of each round, if Shaken, roll one die. On 4+ stop being Shaken",
         "", "", "", "", "", "", "Y", "ROUND_START", "is_shaken", "ROLL", "rally", "4+", "SELF", "10", "Base Rule - Morale"),
        ("FEAR", "Fear(X)", "Counts as having caused +X wounds for morale tests in melee",
         "", "", "", "", "Y", "Y", "", "MORALE", "enemy_checking_morale", "ADD", "effective_wounds", "X", "SELF", "10", "Base Rule - Morale"),

        # Attack modifier rules
        ("FURIOUS", "Furious", "When charging, unmodified rolls of 6 to hit count as extra hits",
         "", "", "Y", "", "Y", "", "", "AFTER_HIT", "is_charging", "ADD", "hits", "sixes_rolled", "SELF", "15", "Base Rule - Attack"),
        ("RELENTLESS", "Relentless", "When shooting at targets more than 9\" away, unmodified 6s to hit count as extra hits",
         "", "", "", "Y", "", "", "", "AFTER_HIT", "range_over_9", "ADD", "hits", "sixes_rolled", "SELF", "15", "Base Rule - Attack"),
        ("SURGE", "Surge", "Unmodified rolls of 6 to hit deal 1 extra hit",
         "", "", "", "Y", "Y", "", "", "AFTER_HIT", "unmodified_6_hit", "ADD", "hits", "sixes_rolled", "SELF", "15", "Base Rule - Attack"),
        ("THRUST", "Thrust", "Gets +1 to hit and +1 AP when charging",
         "", "", "Y", "", "Y", "", "", "HIT_ROLL,CALC_AP", "is_charging", "ADD", "quality_mod,ap", "1,1", "SELF", "10", "Base Rule - Attack"),
        ("PREDATOR_FIGHTER", "Predator Fighter", "For each unmodified 6 to hit in melee, roll one additional attack. Does not apply to newly generated attacks",
         "", "", "", "", "Y", "", "", "HIT_ROLL", "unmodified_6_hit", "ROLL", "extra_attacks", "recursive", "SELF", "20", "Base Rule - Attack"),
        ("GOOD_SHOT", "Good Shot", "Gets +1 to hit when shooting",
         "", "", "", "Y", "", "", "", "HIT_ROLL", "phase_is_shooting", "ADD", "quality_mod", "1", "SELF", "10", "Base Rule - Attack"),
        ("BAD_SHOT", "Bad Shot", "Gets -1 to hit when shooting",
         "", "", "", "Y", "", "", "", "HIT_ROLL", "phase_is_shooting", "ADD", "quality_mod", "-1", "SELF", "10", "Base Rule - Attack"),
        ("VERSATILE_ATTACK", "Versatile Attack", "Before attacking, choose to get either +1 AP or +1 to hit",
         "", "", "", "Y", "Y", "", "", "PRE_ATTACK", "always", "CHOOSE", "ap OR quality_mod", "1", "SELF", "5", "Base Rule - Attack"),
        ("POINT_BLANK_SURGE", "Point-Blank Surge", "When shooting at targets within 9\", unmodified 6s to hit deal extra hits",
         "", "", "", "Y", "", "", "", "AFTER_HIT", "range_0_to_9 AND unmodified_6_hit", "ADD", "hits", "sixes_rolled", "SELF", "15", "Base Rule - Attack"),
        ("PIERCING_ASSAULT", "Piercing Assault", "Gets AP(1) on melee attacks when charging",
         "", "", "Y", "", "Y", "", "", "CALC_AP", "is_charging", "ADD", "ap", "1", "SELF", "10", "Base Rule - Attack"),

        # Wound modifier rules
        ("RUPTURE", "Rupture", "Ignores Regeneration, and unmodified 6s to hit that aren't blocked deal 1 extra wound",
         "", "", "", "Y", "Y", "", "", "AFTER_WOUND,REGEN", "unmodified_6_hit", "ADD,NEGATE", "wounds,regeneration", "sixes_rolled,true", "ENEMY", "15", "Base Rule - Wound"),
        ("SHRED", "Shred", "On unmodified defense rolls of 1, this weapon deals 1 extra wound",
         "", "", "", "Y", "Y", "", "", "AFTER_DEF", "defense_rolled_1", "ADD", "wounds", "ones_rolled", "ENEMY", "15", "Base Rule - Wound"),
        ("SMASH", "Smash", "Ignores Regeneration, and against Defense 5+ or 6+, gets Blast(+3)",
         "", "", "", "Y", "Y", "", "", "AFTER_HIT,REGEN", "target_defense_5plus", "ADD,NEGATE", "hits,regeneration", "3,true", "ENEMY", "15", "Base Rule - Wound"),

        # Allocation rules
        ("HERO", "Hero", "This model takes wounds last in its unit",
         "", "", "", "Y", "Y", "", "", "ALLOCATE", "always", "SET", "allocation_priority", "last", "SELF", "1", "Base Rule - Allocation"),
        ("TAKEDOWN", "Takedown", "Pick one enemy model, attacks against it are resolved as if it were a unit of 1",
         "", "", "", "Y", "Y", "", "", "ALLOCATE", "always", "REPLACE", "allocation_target", "chosen_single", "ENEMY", "1", "Base Rule - Allocation"),

        # Combat order rules
        ("COUNTER", "Counter", "Strikes first when charged",
         "", "", "Y", "", "Y", "", "", "COMBAT_ORDER", "is_charged", "SET", "strike_order", "first", "SELF", "1", "Base Rule - Combat"),
        ("HIT_AND_RUN", "Hit & Run", "May retreat after fighting in melee",
         "", "", "", "", "Y", "", "Y", "ROUND_END", "after_melee", "ENABLE", "retreat_option", "true", "SELF", "10", "Base Rule - Combat"),
        ("SELF_DESTRUCT", "Self-Destruct(X)", "If killed in melee, the attacking unit takes X hits. If survives melee, is killed and enemy takes X hits",
         "", "", "", "", "Y", "", "", "ON_DEATH", "model_killed", "TRIGGER", "hits_to_attacker", "X", "ENEMY", "10", "Base Rule - Combat"),

        # Movement/Deployment rules
        ("SCOUT", "Scout", "May be deployed up to 12\" outside of the deployment zone",
         "Y", "", "", "", "", "", "", "DEPLOYMENT", "always", "ADD", "deploy_distance", "12", "SELF", "10", "Base Rule - Movement"),
        ("AMBUSH", "Ambush", "May be deployed anywhere on the battlefield more than 9\" away from enemies",
         "Y", "", "", "", "", "", "", "DEPLOYMENT", "always", "SET", "deploy_type", "reserve_9", "SELF", "10", "Base Rule - Movement"),
        ("FAST", "Fast", "Moves 9\" instead of the standard 6\"",
         "", "Y", "Y", "", "", "", "", "MOVEMENT", "always", "SET", "move_distance", "9", "SELF", "1", "Base Rule - Movement"),
        ("SLOW", "Slow", "Moves 4\" instead of the standard 6\"",
         "", "Y", "Y", "", "", "", "", "MOVEMENT", "always", "SET", "move_distance", "4", "SELF", "1", "Base Rule - Movement"),
        ("AGILE", "Agile", "Moves +1\" when using Advance, and +2\" when using Rush/Charge",
         "", "Y", "Y", "", "", "", "", "MOVEMENT", "always", "ADD", "advance,rush,charge", "1,2,2", "SELF", "10", "Base Rule - Movement"),
        ("FLYING", "Flying", "May move through units and terrain, ignoring movement penalties",
         "", "Y", "Y", "", "", "", "", "MOVEMENT", "always", "NEGATE", "terrain_penalty,model_blocking", "true,true", "SELF", "10", "Base Rule - Movement"),
        ("STRIDER", "Strider", "Ignores difficult terrain penalties when moving",
         "", "Y", "Y", "", "", "", "", "MOVEMENT", "always", "NEGATE", "difficult_terrain", "true", "SELF", "10", "Base Rule - Movement"),
        ("RAPID_CHARGE", "Rapid Charge", "Moves +4\" when charging",
         "", "", "Y", "", "", "", "", "MOVEMENT", "is_charging", "ADD", "charge_distance", "4", "SELF", "10", "Base Rule - Movement"),

        # Defense modifier rules (hit-based)
        ("MELEE_EVASION", "Melee Evasion", "Enemies get -1 to hit this unit in melee",
         "", "", "", "", "Y", "", "", "HIT_ROLL", "phase_is_melee", "ADD", "enemy_quality_mod", "-1", "SELF", "10", "Base Rule - Defense"),
        ("MELEE_SHROUDING", "Melee Shrouding", "Enemies get -1 to hit this unit in melee",
         "", "", "", "", "Y", "", "", "HIT_ROLL", "phase_is_melee", "ADD", "enemy_quality_mod", "-1", "SELF", "10", "Base Rule - Defense"),
        ("RANGED_SHROUDING", "Ranged Shrouding", "Enemies get -1 to hit this unit when shooting",
         "", "", "", "Y", "", "", "", "HIT_ROLL", "phase_is_shooting", "ADD", "enemy_quality_mod", "-1", "SELF", "10", "Base Rule - Defense"),

        # Special rules
        ("UNSTOPPABLE", "Unstoppable", "Ignores Regeneration when attacking, and ignores negative hit modifiers",
         "", "", "", "Y", "Y", "", "", "HIT_ROLL,REGEN", "always", "NEGATE,NEGATE", "negative_hit_mods,enemy_regeneration", "true,true", "SELF", "5", "Base Rule - Special"),
        ("CASTING", "Casting(X)", "May cast X spells per round",
         "", "", "", "", "", "", "Y", "SPELL_PHASE", "always", "SET", "spell_slots", "X", "SELF", "1", "Base Rule - Magic"),
        ("DEVOUT", "Devout", "Faction-specific bonus (varies by army)",
         "", "", "", "", "", "", "", "VARIES", "faction_specific", "VARIES", "varies", "varies", "SELF", "10", "Base Rule - Faction"),
    ]

def create_rules_matrix():
    wb = Workbook()
    ws = wb.active
    ws.title = "Special Rules"

    # Define headers
    headers = [
        "Rule ID",           # A
        "Rule Name",         # B
        "Description",       # C
        "DEP",               # D
        "MOV",               # E
        "CHG",               # F
        "SHT",               # G
        "MEL",               # H
        "MOR",               # I
        "RND",               # J
        "Trigger Step",      # K
        "Trigger Condition", # L
        "Effect Type",       # M
        "Effect Target",     # N
        "Effect Value",      # O
        "Effect Target Unit",# P
        "Priority",          # Q
        "Notes",             # R
    ]

    # Get base rules
    base_rules = get_base_rules()

    # Extract rules from faction file
    rules_data, armies_per_rule = extract_rules_from_faction_file()

    # Track which base rule names we have (to avoid duplicates from faction file)
    base_rule_names = {rule[1].lower() for rule in base_rules}  # rule[1] is the Rule Name

    # Start with base rules
    processed_rules = list(base_rules)

    # Process faction rules (skip duplicates of base rules)
    for name in sorted(rules_data.keys()):
        # Skip if this is a base rule (case insensitive match)
        if name.lower() in base_rule_names:
            continue
        info = rules_data[name]
        parsed = parse_rule_for_engine(name, info)

        # Truncate description if too long
        desc = info['text']
        if len(desc) > 500:
            desc = desc[:497] + '...'

        # Create notes with army list
        armies = armies_per_rule.get(name, set())
        armies_str = ', '.join(sorted(armies)[:5])  # First 5 armies
        if len(armies) > 5:
            armies_str += f' (+{len(armies)-5} more)'
        notes = f"Type: {info['type']}. Armies: {armies_str}"

        rule_tuple = (
            create_rule_id(name),
            name,
            desc,
            parsed['phases']['DEP'],
            parsed['phases']['MOV'],
            parsed['phases']['CHG'],
            parsed['phases']['SHT'],
            parsed['phases']['MEL'],
            parsed['phases']['MOR'],
            parsed['phases']['RND'],
            parsed['trigger_step'],
            parsed['trigger_condition'],
            parsed['effect_type'],
            parsed['effect_target'],
            parsed['effect_value'],
            parsed['target_unit'],
            parsed['priority'],
            notes
        )
        processed_rules.append(rule_tuple)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    phase_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    engine_header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")

    y_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

        if col <= 3:
            cell.fill = header_fill
        elif col <= 10:
            cell.fill = phase_header_fill
        else:
            cell.fill = engine_header_fill

    # Write data
    for row_idx, rule in enumerate(processed_rules, 2):
        for col_idx, value in enumerate(rule, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            if 4 <= col_idx <= 10:
                cell.alignment = center_align
                if value == "Y":
                    cell.fill = y_fill
                    cell.font = Font(bold=True)
            elif col_idx in [3, 18]:
                cell.alignment = wrap_align
            elif col_idx in [1, 2, 13, 16, 17]:
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Column widths
    column_widths = {
        'A': 25, 'B': 30, 'C': 60,
        'D': 5, 'E': 5, 'F': 5, 'G': 5, 'H': 5, 'I': 5, 'J': 5,
        'K': 18, 'L': 30, 'M': 12, 'N': 20, 'O': 20, 'P': 12, 'Q': 8, 'R': 50,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(2, len(processed_rules) + 2):
        ws.row_dimensions[row].height = 40

    ws.freeze_panes = 'D2'

    # Data validation
    effect_types = '"ADD,MULTIPLY,SET,REROLL,REPLACE,NEGATE,ROLL,CHOOSE,TRIGGER,ENABLE,GRANT,DAMAGE"'
    dv_effect = DataValidation(type="list", formula1=effect_types, allow_blank=True)
    ws.add_data_validation(dv_effect)
    dv_effect.add(f'M2:M{len(processed_rules)+1}')

    target_units = '"SELF,ENEMY,WEAPON,FRIENDLY"'
    dv_target = DataValidation(type="list", formula1=target_units, allow_blank=True)
    ws.add_data_validation(dv_target)
    dv_target.add(f'P2:P{len(processed_rules)+1}')

    # Legend sheet
    ws_legend = wb.create_sheet("Legend")
    legend_data = [
        ("COLUMN REFERENCE", "", ""),
        ("", "", ""),
        ("Column", "Values", "Description"),
        ("", "", ""),
        ("Rule ID", "(text)", "Unique identifier used in code"),
        ("Rule Name", "(text)", "Display name with parameter notation"),
        ("Description", "(text)", "Full rule text from army book"),
        ("", "", ""),
        ("PHASE COLUMNS", "", "Mark Y if rule applies in this phase"),
        ("DEP", "Y or blank", "Deployment - before battle begins"),
        ("MOV", "Y or blank", "Movement phase"),
        ("CHG", "Y or blank", "Charge declaration and movement"),
        ("SHT", "Y or blank", "Shooting phase"),
        ("MEL", "Y or blank", "Melee combat phase"),
        ("MOR", "Y or blank", "Morale checks"),
        ("RND", "Y or blank", "Round start/end effects"),
        ("", "", ""),
        ("ENGINE CONFIGURATION", "", ""),
        ("Trigger Step", "(see list)", "When in combat resolution the rule activates"),
        ("", "CALC_ATTACKS", "When determining number of attacks"),
        ("", "HIT_ROLL", "During the quality/hit test"),
        ("", "AFTER_HIT", "After hits determined, before defense"),
        ("", "CALC_AP", "When calculating armor penetration"),
        ("", "DEF_ROLL", "During the defense/save test"),
        ("", "AFTER_DEF", "After defense roll resolved"),
        ("", "AFTER_WOUND", "After wounds calculated"),
        ("", "ALLOCATE", "When assigning wounds to models"),
        ("", "REGEN", "During regeneration rolls"),
        ("", "MORALE", "During morale tests"),
        ("", "ROUND_START", "At beginning of round"),
        ("", "ROUND_END", "At end of round"),
        ("", "ON_DEATH", "When a model is removed"),
        ("", "PRE_ATTACK", "Before attack sequence begins"),
        ("", "COMBAT_ORDER", "When determining strike order"),
        ("", "MOVEMENT", "During movement"),
        ("", "DEPLOYMENT", "During deployment"),
        ("", "DIRECT_DAMAGE", "Direct damage outside normal attacks"),
        ("", "", ""),
        ("Trigger Condition", "(expression)", "Boolean condition that must be true"),
        ("", "always", "Always triggers"),
        ("", "is_charging", "Unit is charging this round"),
        ("", "is_charged", "Unit was charged this round"),
        ("", "unmodified_6_hit", "Natural 6 rolled on hit"),
        ("", "phase_is_shooting", "Currently in shooting phase"),
        ("", "phase_is_melee", "Currently in melee phase"),
        ("", "range_over_9", "Target is more than 9\" away"),
        ("", "range_over_12", "Target is more than 12\" away"),
        ("", "attack_not_spell", "Attack is not a spell"),
        ("", "once_per_game", "Can only be used once per game"),
        ("", "once_per_activation", "Can only be used once per activation"),
        ("", "", ""),
        ("Effect Type", "(from list)", "How the rule modifies values"),
        ("", "ADD", "Add value to target (can be negative)"),
        ("", "MULTIPLY", "Multiply target by value"),
        ("", "SET", "Replace target with value"),
        ("", "REROLL", "Allow reroll of specified dice"),
        ("", "REPLACE", "Replace behavior entirely"),
        ("", "NEGATE", "Cancel/ignore something"),
        ("", "ROLL", "Require a dice roll for effect"),
        ("", "CHOOSE", "Player/AI chooses between options"),
        ("", "TRIGGER", "Cause a secondary effect"),
        ("", "ENABLE", "Unlock an optional action"),
        ("", "GRANT", "Grant a rule to another unit"),
        ("", "DAMAGE", "Deal direct damage"),
        ("", "", ""),
        ("Effect Target", "(stat name)", "What is being modified"),
        ("", "attacks", "Number of attack dice"),
        ("", "hits", "Number of successful hits"),
        ("", "quality_mod", "Modifier to quality roll"),
        ("", "ap", "Armor penetration value"),
        ("", "defense_target", "Defense roll target number"),
        ("", "defense_mod", "Modifier to defense roll"),
        ("", "wounds", "Number of wounds dealt"),
        ("", "regeneration", "Regeneration ability"),
        ("", "morale_mod", "Modifier to morale test"),
        ("", "direct_hits", "Direct hit damage"),
        ("", "", ""),
        ("Target Unit", "(from list)", "Who is affected"),
        ("", "SELF", "The unit with this rule"),
        ("", "ENEMY", "The opposing unit"),
        ("", "WEAPON", "The weapon being used"),
        ("", "FRIENDLY", "A friendly unit (for auras)"),
        ("", "", ""),
        ("Priority", "(number)", "Lower numbers process first"),
    ]

    for row_idx, (col1, col2, col3) in enumerate(legend_data, 1):
        ws_legend.cell(row=row_idx, column=1, value=col1)
        ws_legend.cell(row=row_idx, column=2, value=col2)
        ws_legend.cell(row=row_idx, column=3, value=col3)

        if col1 in ["COLUMN REFERENCE", "PHASE COLUMNS", "ENGINE CONFIGURATION"] or (col1 == "Column"):
            ws_legend.cell(row=row_idx, column=1).font = Font(bold=True)
            if col1 == "Column":
                ws_legend.cell(row=row_idx, column=2).font = Font(bold=True)
                ws_legend.cell(row=row_idx, column=3).font = Font(bold=True)
        elif col1 and col2:
            ws_legend.cell(row=row_idx, column=1).font = Font(bold=True)

    ws_legend.column_dimensions['A'].width = 20
    ws_legend.column_dimensions['B'].width = 25
    ws_legend.column_dimensions['C'].width = 50

    # Save
    output_path = "/home/user/Science-Battle-Simulator/docs/special_rules_review_matrix.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    print(f"Total rules: {len(processed_rules)}")
    return output_path

if __name__ == "__main__":
    create_rules_matrix()
