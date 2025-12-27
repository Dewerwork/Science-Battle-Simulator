#!/usr/bin/env python3
"""
Validation script for merged loadout files.
Compares source file against reduced file and generates an Excel report.

Supports both JSON (from merge_all_factions.py) and TXT formats.
When reading JSON, properly extracts unit lines from the factions structure.
"""

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    exit(1)

# =============================================================================
# CONFIGURATION - Hardcoded file paths
# =============================================================================

SOURCE_FILE = r"C:\Users\David\Documents\Army Factions\pipeline_output\all_factions_merged.txt"
REDUCED_FILE = r"C:\Users\David\Documents\Army Factions\pipeline_output\all_faction.reduced.txt"
OUTPUT_WORKBOOK = r"C:\Users\David\Documents\Army Factions\pipeline_output\validation_report.xlsx"

# =============================================================================
# Data Structures
# =============================================================================

@dataclass
class UnitEntry:
    """Represents a single unit loadout entry from the merged file."""
    raw_header: str
    base_name: str              # Unit name without UID/BKT/FID
    identifier: str             # UID or BKT hash (empty if none)
    identifier_type: str        # 'UID', 'BKT', or ''
    faction_id: str             # FID hash (empty if none) - kept separate for grouping
    size: int
    quality: int
    defense: int
    points: int
    rules: List[str]
    weapon_lines: List[str]
    line_number: int
    parse_errors: List[str] = field(default_factory=list)


@dataclass
class ValidationResult:
    """Result of a single validation test."""
    test_name: str
    passed: bool
    pass_count: int
    fail_count: int
    severity: str               # Critical, Warning, Info
    failures: List[dict]        # Details for sheet output


# =============================================================================
# Parsing Functions
# =============================================================================

# Header pattern: UnitName [UID:XXXX] [size] Q#+ D#+ | points pts | rules
HEADER_PATTERN = re.compile(
    r'^(.+?)\s+'                           # Unit name (non-greedy)
    r'\[(\d+)\]\s+'                         # Size
    r'Q(\d)\+\s+D(\d)\+\s*'                 # Quality and Defense
    r'\|\s*(\d+)pts\s*\|'                   # Points
    r'\s*(.*)$'                             # Rules (rest of line)
)

# Pattern to extract UID, BKT, or FID from unit name
IDENTIFIER_PATTERN = re.compile(r'\[(UID|BKT|FID):([A-F0-9]+)\]', re.IGNORECASE)

# Weapon line patterns
WEAPON_PATTERN = re.compile(r'(\d+x\s+)?(\d+")?\s*(.+?)\s*\(([^)]+)\)')


def extract_identifier(name: str) -> Tuple[str, str, str, str]:
    """
    Extract identifier and faction ID from unit name.
    Returns: (base_name, identifier, identifier_type, faction_id)

    FID is extracted separately since it's used for grouping units by faction,
    while UID/BKT are used for other purposes.
    """
    faction_id = ''
    identifier = ''
    identifier_type = ''

    # Extract all identifiers
    for match in IDENTIFIER_PATTERN.finditer(name):
        id_type = match.group(1).upper()
        id_value = match.group(2).upper()

        if id_type == 'FID':
            faction_id = id_value
        else:
            # UID or BKT - keep the last one found
            identifier = id_value
            identifier_type = id_type

    # Remove all identifier tags from name
    base_name = IDENTIFIER_PATTERN.sub('', name).strip()
    return base_name, identifier, identifier_type, faction_id


def parse_header(header: str, line_number: int) -> Optional[UnitEntry]:
    """Parse a unit header line into a UnitEntry."""
    errors = []

    match = HEADER_PATTERN.match(header.strip())
    if not match:
        # Try to extract what we can
        errors.append(f"Header does not match expected pattern")
        return UnitEntry(
            raw_header=header,
            base_name=header[:50] + "..." if len(header) > 50 else header,
            identifier='',
            identifier_type='',
            faction_id='',
            size=0,
            quality=0,
            defense=0,
            points=0,
            rules=[],
            weapon_lines=[],
            line_number=line_number,
            parse_errors=errors
        )

    name_with_id = match.group(1).strip()
    base_name, identifier, id_type, faction_id = extract_identifier(name_with_id)

    try:
        size = int(match.group(2))
    except ValueError:
        size = 0
        errors.append(f"Invalid size: {match.group(2)}")

    try:
        quality = int(match.group(3))
    except ValueError:
        quality = 0
        errors.append(f"Invalid quality: {match.group(3)}")

    try:
        defense = int(match.group(4))
    except ValueError:
        defense = 0
        errors.append(f"Invalid defense: {match.group(4)}")

    try:
        points = int(match.group(5))
    except ValueError:
        points = 0
        errors.append(f"Invalid points: {match.group(5)}")

    rules_str = match.group(6).strip()
    rules = [r.strip() for r in rules_str.split(',') if r.strip()] if rules_str else []

    return UnitEntry(
        raw_header=header,
        base_name=base_name,
        identifier=identifier,
        identifier_type=id_type,
        faction_id=faction_id,
        size=size,
        quality=quality,
        defense=defense,
        points=points,
        rules=rules,
        weapon_lines=[],
        line_number=line_number,
        parse_errors=errors
    )


def parse_lines_to_entries(lines: List[str], start_line: int = 1) -> List[UnitEntry]:
    """Parse a list of lines into UnitEntry objects."""
    entries = []
    current_entry = None
    line_number = start_line

    for line in lines:
        stripped = line.strip()

        # Skip empty lines - they separate entries
        if not stripped:
            if current_entry:
                entries.append(current_entry)
                current_entry = None
            line_number += 1
            continue

        # Check if this looks like a header line (has Q#+ D#+ pattern)
        if re.search(r'Q\d\+\s+D\d\+', stripped):
            if current_entry:
                entries.append(current_entry)
            current_entry = parse_header(stripped, line_number)
        elif current_entry:
            # This is a weapon line
            current_entry.weapon_lines.append(stripped)

        line_number += 1

    # Don't forget the last entry
    if current_entry:
        entries.append(current_entry)

    return entries


def parse_merged_file(filepath: str) -> List[UnitEntry]:
    """Parse a merged TXT or JSON file into a list of UnitEntry objects.

    Detects JSON content by checking both file extension and file content,
    so JSON files with .txt extension are handled correctly.
    """
    entries = []

    path = Path(filepath)
    if not path.exists():
        print(f"ERROR: File not found: {filepath}")
        return entries

    # Read file content first
    with open(path, 'r', encoding='utf-8-sig') as f:
        content = f.read()

    # Detect JSON by extension OR by content starting with '{'
    is_json = path.suffix.lower() == '.json' or content.lstrip().startswith('{')

    if is_json:
        try:
            data = json.loads(content)

            # Parse JSON format from merge_all_factions.py
            line_number = 1
            for faction in data.get("factions", []):
                faction_name = faction.get("name", "Unknown")
                unit_lines = faction.get("units", [])
                faction_entries = parse_lines_to_entries(unit_lines, line_number)
                entries.extend(faction_entries)
                line_number += len(unit_lines) + 1
                print(f"    {faction_name}: {len(faction_entries)} entries")

            return entries
        except json.JSONDecodeError as e:
            print(f"ERROR: Failed to parse JSON: {e}")
            print("  Falling back to text parsing...")

    # Parse as TXT file
    lines = content.splitlines()
    return parse_lines_to_entries(lines)


def _legacy_parse_merged_file(filepath: str) -> List[UnitEntry]:
    """Legacy parser - kept for reference."""
    entries = []

    path = Path(filepath)
    if not path.exists():
        print(f"ERROR: File not found: {filepath}")
        return entries

    with open(path, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    current_entry = None
    line_number = 0

    for line in lines:
        line_number += 1
        stripped = line.strip()

        # Skip empty lines - they separate entries
        if not stripped:
            if current_entry:
                entries.append(current_entry)
                current_entry = None
            continue

        # Check if this looks like a header line (has Q#+ D#+ pattern)
        if re.search(r'Q\d\+\s+D\d\+', stripped):
            if current_entry:
                entries.append(current_entry)
            current_entry = parse_header(stripped, line_number)
        elif current_entry:
            # This is a weapon line
            current_entry.weapon_lines.append(stripped)

    # Don't forget the last entry
    if current_entry:
        entries.append(current_entry)

    return entries


# =============================================================================
# Validation Functions
# =============================================================================

def validate_completeness(source: List[UnitEntry], reduced: List[UnitEntry]) -> ValidationResult:
    """Check that all units in source appear in reduced."""
    source_names = defaultdict(int)
    for entry in source:
        source_names[entry.base_name] += 1

    reduced_names = defaultdict(int)
    for entry in reduced:
        reduced_names[entry.base_name] += 1

    missing = []
    for name, count in source_names.items():
        if name not in reduced_names:
            missing.append({
                'Unit Name': name,
                'Source Count': count,
                'Reduced Count': 0,
                'Issue': 'Missing from reduced file'
            })

    return ValidationResult(
        test_name='Completeness Check',
        passed=len(missing) == 0,
        pass_count=len(source_names) - len(missing),
        fail_count=len(missing),
        severity='Critical',
        failures=missing
    )


def validate_orphans(source: List[UnitEntry], reduced: List[UnitEntry]) -> ValidationResult:
    """Check for units in reduced that don't exist in source."""
    source_names = set(entry.base_name for entry in source)

    reduced_counts = defaultdict(int)
    for entry in reduced:
        reduced_counts[entry.base_name] += 1

    orphans = []
    for name, count in reduced_counts.items():
        if name not in source_names:
            orphans.append({
                'Unit Name': name,
                'Reduced Count': count,
                'Issue': 'Not found in source file (orphaned)'
            })

    return ValidationResult(
        test_name='Orphan Detection',
        passed=len(orphans) == 0,
        pass_count=len(reduced_counts) - len(orphans),
        fail_count=len(orphans),
        severity='Critical',
        failures=orphans
    )


def validate_header_format(entries: List[UnitEntry], file_label: str) -> ValidationResult:
    """Validate header format for all entries."""
    failures = []

    for entry in entries:
        if entry.parse_errors:
            failures.append({
                'Line Number': entry.line_number,
                'Unit Name': entry.base_name[:50],
                'ID': f"{entry.identifier_type}:{entry.identifier}" if entry.identifier else '',
                'Raw Header': entry.raw_header[:100] + ('...' if len(entry.raw_header) > 100 else ''),
                'Issues': '; '.join(entry.parse_errors),
                'File': file_label
            })

    return ValidationResult(
        test_name='Header Format',
        passed=len(failures) == 0,
        pass_count=len(entries) - len(failures),
        fail_count=len(failures),
        severity='Warning',
        failures=failures
    )


def validate_weapon_format(entries: List[UnitEntry], file_label: str) -> ValidationResult:
    """Validate weapon line format."""
    failures = []

    # Known special rules that should NOT be weapon names
    rule_names = {'blast', 'deadly', 'rending', 'poison', 'reliable', 'sniper', 'impact', 'furious'}

    for entry in entries:
        for weapon_line in entry.weapon_lines:
            issues = []
            line_stripped = weapon_line.strip()

            # Check for empty/placeholder weapon lines
            if line_stripped == '-' or line_stripped == '':
                issues.append('Empty/placeholder weapon line')

            # Check for zero-attack weapons (A0)
            if re.search(r'\(A0[,\)]', weapon_line) or re.search(r'\(A0\s', weapon_line):
                issues.append('Zero-attack weapon (A0)')

            # Check for unbalanced parentheses
            open_count = weapon_line.count('(')
            close_count = weapon_line.count(')')
            if open_count != close_count:
                issues.append(f'Unbalanced parentheses ({open_count} open, {close_count} close)')

            # Check for rule names used as weapon names (first word before parentheses)
            first_word_match = re.match(r'^(\d+x\s+)?(\d+")?\s*(\w+)', line_stripped)
            if first_word_match:
                potential_name = first_word_match.group(3).lower() if first_word_match.group(3) else ''
                if potential_name in rule_names:
                    issues.append(f'Rule name "{potential_name}" used as weapon name')

            # Check for missing weapon profile parentheses (only if not already flagged)
            if not issues and not re.search(r'\([^)]+\)', weapon_line):
                issues.append('Missing weapon profile parentheses')

            if issues:
                failures.append({
                    'Line Number': entry.line_number,
                    'Unit Name': entry.base_name[:40],
                    'ID': f"{entry.identifier_type}:{entry.identifier}" if entry.identifier else '',
                    'Weapon Line': weapon_line[:80] + ('...' if len(weapon_line) > 80 else ''),
                    'Issues': '; '.join(issues),
                    'File': file_label
                })

    return ValidationResult(
        test_name='Weapon Format',
        passed=len(failures) == 0,
        pass_count=sum(len(e.weapon_lines) for e in entries) - len(failures),
        fail_count=len(failures),
        severity='Warning',
        failures=failures
    )


def validate_stats_consistency(entries: List[UnitEntry], file_label: str) -> ValidationResult:
    """Check for units with inconsistent Q/D stats across entries.

    Groups units by base_name + faction_id, so units with the same name
    from different factions are treated as separate units.
    """
    # Group by base name + faction ID and check for stat variations
    # Key is (base_name, faction_id) tuple
    unit_stats: Dict[Tuple[str, str], List[Tuple[int, int, int]]] = defaultdict(list)

    for entry in entries:
        if entry.quality > 0 and entry.defense > 0:  # Only valid entries
            key = (entry.base_name, entry.faction_id)
            unit_stats[key].append((entry.quality, entry.defense, entry.line_number))

    failures = []
    for (name, fid), stats_list in unit_stats.items():
        unique_stats = set((q, d) for q, d, _ in stats_list)
        if len(unique_stats) > 1:
            stats_str = ', '.join(f"Q{q}+ D{d}+" for q, d in sorted(unique_stats))
            # Include FID in the output if present
            display_name = f"{name} [FID:{fid}]" if fid else name
            failures.append({
                'Unit Name': display_name,
                'Variations Found': stats_str,
                'Entry Count': len(stats_list),
                'Issue': 'Inconsistent Q/D stats across entries',
                'File': file_label
            })

    return ValidationResult(
        test_name='Stats Consistency',
        passed=len(failures) == 0,
        pass_count=len(unit_stats) - len(failures),
        fail_count=len(failures),
        severity='Warning',
        failures=failures
    )


def validate_duplicate_buckets(entries: List[UnitEntry]) -> ValidationResult:
    """Check for duplicate bucket hashes within the same unit."""
    # Only applies to reduced file with BKT identifiers
    unit_buckets: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for entry in entries:
        if entry.identifier_type == 'BKT' and entry.identifier:
            unit_buckets[entry.base_name][entry.identifier] += 1

    failures = []
    for name, buckets in unit_buckets.items():
        for bucket, count in buckets.items():
            if count > 1:
                failures.append({
                    'Unit Name': name,
                    'ID': f"BKT:{bucket}",
                    'Occurrences': count,
                    'Issue': 'Duplicate bucket hash for same unit'
                })

    return ValidationResult(
        test_name='Duplicate Buckets',
        passed=len(failures) == 0,
        pass_count=sum(len(b) for b in unit_buckets.values()) - len(failures),
        fail_count=len(failures),
        severity='Warning',
        failures=failures
    )


def validate_rules(entries: List[UnitEntry], file_label: str) -> ValidationResult:
    """Check for malformed or suspicious rules."""
    failures = []

    # Patterns for suspicious rules
    cost_pattern = re.compile(r'[+-]\d+pts?', re.IGNORECASE)
    truncated_patterns = [
        (re.compile(r'^casting\s+\w{1,4}$', re.IGNORECASE), 'Truncated Casting rule'),
        (re.compile(r'^tough\s*$', re.IGNORECASE), 'Incomplete Tough rule'),
        (re.compile(r'^\w+\(\s*$'), 'Incomplete rule with open paren'),
    ]
    weapon_leak_pattern = re.compile(r'^A\d+,\s*AP\(\d+\)')

    for entry in entries:
        for rule in entry.rules:
            issues = []

            # Check for cost artifacts
            if cost_pattern.search(rule):
                issues.append('Contains cost modifier')

            # Check for truncated rules
            for pattern, desc in truncated_patterns:
                if pattern.match(rule):
                    issues.append(desc)
                    break

            # Check for weapon profile leaking into rules
            if weapon_leak_pattern.match(rule):
                issues.append('Weapon profile in rules')

            # Check for numeric-only rules (model count leaking into rules)
            if re.match(r'^\d+$', rule.strip()):
                issues.append('Numeric-only rule (likely model count leak)')

            # Check for very short rules (likely truncated)
            if len(rule) < 3 and not rule.isdigit():
                issues.append('Suspiciously short rule')

            if issues:
                failures.append({
                    'Line Number': entry.line_number,
                    'Unit Name': entry.base_name[:40],
                    'ID': f"{entry.identifier_type}:{entry.identifier}" if entry.identifier else '',
                    'Rule': rule,
                    'Issues': '; '.join(issues),
                    'File': file_label
                })

    total_rules = sum(len(e.rules) for e in entries)
    return ValidationResult(
        test_name='Rule Issues',
        passed=len(failures) == 0,
        pass_count=total_rules - len(failures),
        fail_count=len(failures),
        severity='Warning',
        failures=failures
    )


def validate_edge_cases(entries: List[UnitEntry], file_label: str) -> ValidationResult:
    """Check for edge cases: zero points, extreme values, empty weapons."""
    failures = []

    for entry in entries:
        issues = []

        # Zero points
        if entry.points == 0:
            issues.append('Zero points')

        # Very low points
        if 0 < entry.points < 5:
            issues.append(f'Very low points ({entry.points})')

        # Empty weapons
        if not entry.weapon_lines:
            issues.append('No weapons')

        # Zero size
        if entry.size == 0:
            issues.append('Zero unit size')

        # Very large size
        if entry.size > 30:
            issues.append(f'Very large unit size ({entry.size})')

        if issues:
            failures.append({
                'Line Number': entry.line_number,
                'Unit Name': entry.base_name[:40],
                'ID': f"{entry.identifier_type}:{entry.identifier}" if entry.identifier else '',
                'Points': entry.points,
                'Size': entry.size,
                'Issues': '; '.join(issues),
                'File': file_label
            })

    return ValidationResult(
        test_name='Edge Cases',
        passed=len(failures) == 0,
        pass_count=len(entries) - len(failures),
        fail_count=len(failures),
        severity='Info',
        failures=failures
    )


# =============================================================================
# Excel Report Generation
# =============================================================================

# Color definitions
COLORS = {
    'pass': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),      # Light green
    'fail': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),      # Light red
    'warning': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),   # Light yellow
    'header': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),    # Blue
    'info': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),      # Light blue
}

HEADER_FONT = Font(bold=True, color='FFFFFF')
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def auto_width(ws, min_width=10, max_width=60):
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = None
        for cell in column_cells:
            # Skip merged cells which don't have column_letter
            if hasattr(cell, 'column_letter'):
                if column is None:
                    column = cell.column_letter
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_len)
                except:
                    pass
        if column:
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width


def write_summary_sheet(ws, results: List[ValidationResult], source_count: int, reduced_count: int):
    """Write the summary sheet."""
    # Title and metadata
    ws['A1'] = 'Validation Report'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'] = f'Source File: {SOURCE_FILE}'
    ws['A4'] = f'Reduced File: {REDUCED_FILE}'
    ws['A5'] = f'Source Entries: {source_count}'
    ws['A6'] = f'Reduced Entries: {reduced_count}'

    # Summary table header
    headers = ['Test Name', 'Status', 'Pass Count', 'Fail Count', 'Severity']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.fill = COLORS['header']
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Summary data
    for row, result in enumerate(results, 9):
        ws.cell(row=row, column=1, value=result.test_name).border = THIN_BORDER

        status_cell = ws.cell(row=row, column=2, value='PASS' if result.passed else 'FAIL')
        status_cell.border = THIN_BORDER
        status_cell.alignment = Alignment(horizontal='center')
        if result.passed:
            status_cell.fill = COLORS['pass']
        elif result.severity == 'Critical':
            status_cell.fill = COLORS['fail']
        else:
            status_cell.fill = COLORS['warning']

        ws.cell(row=row, column=3, value=result.pass_count).border = THIN_BORDER
        ws.cell(row=row, column=4, value=result.fail_count).border = THIN_BORDER

        severity_cell = ws.cell(row=row, column=5, value=result.severity)
        severity_cell.border = THIN_BORDER
        severity_cell.alignment = Alignment(horizontal='center')
        if result.severity == 'Critical':
            severity_cell.fill = COLORS['fail']
        elif result.severity == 'Warning':
            severity_cell.fill = COLORS['warning']
        else:
            severity_cell.fill = COLORS['info']

    auto_width(ws)


def write_detail_sheet(ws, result: ValidationResult):
    """Write a detail sheet for a validation result."""
    # Excel max rows is 1,048,576 - leave room for header and truncation warning
    MAX_DETAIL_ROWS = 1_000_000

    if not result.failures:
        ws['A1'] = 'No issues found - all checks passed!'
        ws['A1'].fill = COLORS['pass']
        ws['A1'].font = BOLD_FONT
        return

    # Get headers from first failure
    headers = list(result.failures[0].keys())

    # Check if we need to truncate
    truncated = len(result.failures) > MAX_DETAIL_ROWS
    failures_to_write = result.failures[:MAX_DETAIL_ROWS] if truncated else result.failures

    # Write truncation warning if needed
    start_row = 1
    if truncated:
        ws['A1'] = f'WARNING: Results truncated! Showing {MAX_DETAIL_ROWS:,} of {len(result.failures):,} failures.'
        ws['A1'].fill = COLORS['warning']
        ws['A1'].font = BOLD_FONT
        start_row = 3

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = COLORS['header']
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for row, failure in enumerate(failures_to_write, start_row + 1):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=failure.get(header, ''))
            cell.border = THIN_BORDER

    auto_width(ws)


def generate_report(results: List[ValidationResult], source_count: int, reduced_count: int):
    """Generate the Excel workbook."""
    wb = Workbook()

    # Summary sheet (rename default sheet)
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    write_summary_sheet(ws_summary, results, source_count, reduced_count)

    # Detail sheets
    sheet_names = [
        'Completeness',
        'Orphans',
        'Header Format',
        'Weapon Format',
        'Stats Mismatch',
        'Duplicate Buckets',
        'Rule Issues',
        'Edge Cases'
    ]

    for result, sheet_name in zip(results, sheet_names):
        ws = wb.create_sheet(title=sheet_name)
        write_detail_sheet(ws, result)

    # Save workbook
    output_path = Path(OUTPUT_WORKBOOK)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\nReport saved to: {output_path}")


# =============================================================================
# Main
# =============================================================================

def main():
    print("=" * 60)
    print("MERGED LOADOUT VALIDATION")
    print("=" * 60)

    # Parse files
    print(f"\nLoading source file: {SOURCE_FILE}")
    source_entries = parse_merged_file(SOURCE_FILE)
    print(f"  Found {len(source_entries)} entries")

    print(f"\nLoading reduced file: {REDUCED_FILE}")
    reduced_entries = parse_merged_file(REDUCED_FILE)
    print(f"  Found {len(reduced_entries)} entries")

    if not source_entries:
        print("\nERROR: No entries found in source file. Aborting.")
        return 1

    if not reduced_entries:
        print("\nERROR: No entries found in reduced file. Aborting.")
        return 1

    # Run validations
    print("\nRunning validations...")
    results = []

    # 1. Completeness
    print("  [1/8] Completeness check...")
    results.append(validate_completeness(source_entries, reduced_entries))

    # 2. Orphans
    print("  [2/8] Orphan detection...")
    results.append(validate_orphans(source_entries, reduced_entries))

    # 3. Header format (combined from both files)
    print("  [3/8] Header format validation...")
    source_header_result = validate_header_format(source_entries, 'Source')
    reduced_header_result = validate_header_format(reduced_entries, 'Reduced')
    combined_header = ValidationResult(
        test_name='Header Format',
        passed=source_header_result.passed and reduced_header_result.passed,
        pass_count=source_header_result.pass_count + reduced_header_result.pass_count,
        fail_count=source_header_result.fail_count + reduced_header_result.fail_count,
        severity='Warning',
        failures=source_header_result.failures + reduced_header_result.failures
    )
    results.append(combined_header)

    # 4. Weapon format
    print("  [4/8] Weapon format validation...")
    source_weapon_result = validate_weapon_format(source_entries, 'Source')
    reduced_weapon_result = validate_weapon_format(reduced_entries, 'Reduced')
    combined_weapon = ValidationResult(
        test_name='Weapon Format',
        passed=source_weapon_result.passed and reduced_weapon_result.passed,
        pass_count=source_weapon_result.pass_count + reduced_weapon_result.pass_count,
        fail_count=source_weapon_result.fail_count + reduced_weapon_result.fail_count,
        severity='Warning',
        failures=source_weapon_result.failures + reduced_weapon_result.failures
    )
    results.append(combined_weapon)

    # 5. Stats consistency
    print("  [5/8] Stats consistency check...")
    source_stats_result = validate_stats_consistency(source_entries, 'Source')
    reduced_stats_result = validate_stats_consistency(reduced_entries, 'Reduced')
    combined_stats = ValidationResult(
        test_name='Stats Consistency',
        passed=source_stats_result.passed and reduced_stats_result.passed,
        pass_count=source_stats_result.pass_count + reduced_stats_result.pass_count,
        fail_count=source_stats_result.fail_count + reduced_stats_result.fail_count,
        severity='Warning',
        failures=source_stats_result.failures + reduced_stats_result.failures
    )
    results.append(combined_stats)

    # 6. Duplicate buckets (reduced file only)
    print("  [6/8] Duplicate bucket detection...")
    results.append(validate_duplicate_buckets(reduced_entries))

    # 7. Rule issues
    print("  [7/8] Rule validation...")
    source_rules_result = validate_rules(source_entries, 'Source')
    reduced_rules_result = validate_rules(reduced_entries, 'Reduced')
    combined_rules = ValidationResult(
        test_name='Rule Issues',
        passed=source_rules_result.passed and reduced_rules_result.passed,
        pass_count=source_rules_result.pass_count + reduced_rules_result.pass_count,
        fail_count=source_rules_result.fail_count + reduced_rules_result.fail_count,
        severity='Warning',
        failures=source_rules_result.failures + reduced_rules_result.failures
    )
    results.append(combined_rules)

    # 8. Edge cases
    print("  [8/8] Edge case detection...")
    source_edge_result = validate_edge_cases(source_entries, 'Source')
    reduced_edge_result = validate_edge_cases(reduced_entries, 'Reduced')
    combined_edge = ValidationResult(
        test_name='Edge Cases',
        passed=source_edge_result.passed and reduced_edge_result.passed,
        pass_count=source_edge_result.pass_count + reduced_edge_result.pass_count,
        fail_count=source_edge_result.fail_count + reduced_edge_result.fail_count,
        severity='Info',
        failures=source_edge_result.failures + reduced_edge_result.failures
    )
    results.append(combined_edge)

    # Generate report
    print("\nGenerating Excel report...")
    generate_report(results, len(source_entries), len(reduced_entries))

    # Console summary
    print("\n" + "=" * 60)
    print("VALIDATION SUMMARY")
    print("=" * 60)

    all_passed = True
    for result in results:
        status = "PASS" if result.passed else "FAIL"
        icon = "✓" if result.passed else "✗"
        print(f"  [{status}] {result.test_name}: {result.pass_count} passed, {result.fail_count} failed")
        if not result.passed:
            all_passed = False

    print("=" * 60)
    if all_passed:
        print("All validations PASSED!")
    else:
        print("Some validations FAILED - see Excel report for details.")

    return 0 if all_passed else 1


if __name__ == '__main__':
    exit(main())
