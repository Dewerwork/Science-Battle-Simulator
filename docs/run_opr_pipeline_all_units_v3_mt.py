from __future__ import annotations

import hashlib
import itertools
import json
import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from concurrent.futures import ProcessPoolExecutor, as_completed

# =========================================================
# HARD-CODED SETTINGS (edit these in PyCharm)
# =========================================================
PDF_PATH = r"C:\Users\David\Documents\Army Factions\GF - Blessed Sisters 3.5.1.pdf"
RULES_XLSX_PATH = r"C:\Users\David\Desktop\Special Rules Exclusion List.xlsx"
OUTPUT_DIR = r"C:\Users\David\Documents\Army Factions\Blessed_Sisters_pipeline"

FACTION_NAME = "Blessed Sisters"
FACTION_VERSION = "3.5.1"

# 0 = no limit (can explode!)
MAX_LOADOUTS_PER_UNIT = 0

# Within-unit parallelism (processes, not threads)
WORKERS_PER_UNIT = 32         # set 24 or 32 for your i9
TASKS_PER_UNIT = 256 # good load balancing for uneven work

# Write huge ungrouped loadouts? (usually NO)
WRITE_UNGROUPED_LOADOUTS_TXT = False

# Stage-1 rule policy driven by XLSX
RULE_POLICY = "exclude"   # "exclude" | "bucket" | "keep"
INCLUDE_POINTS_IN_STAGE1_SIGNATURE = True  # stage-2 ignores points anyway, but can help lineage/debug

# Stage-2 weapon condensation settings
RANGE_BUCKETS = [6, 12, 18, 24]     # anything above -> "32+"
RANGE_BUCKET_HIGH = "32+"

# TXT formatting
ADD_BLANK_LINE_BETWEEN_UNITS = True

# =========================================================
# Utilities
# =========================================================
def sha1(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()

def safe_filename(name: str) -> str:
    name = name.strip()
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_") or "unit"

def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)

def norm_ws(s: str) -> str:
    return " ".join(str(s).strip().split())

# =========================================================
# 1) PDF -> units JSON
# =========================================================
UNIT_HEADER_RE = re.compile(r"^(?P<name>.+?)\s+\[(?P<size>\d+)\]\s*-\s*(?P<pts>\d+)pts$")
STATS_RE = re.compile(r"^Quality\s+(?P<qua>\d)\+\s+Defense\s+(?P<def>\d)\+(?:\s+Tough\s+(?P<tough>\d+))?$")
WEAPON_HEADER = "Weapon RNG ATK AP SPE"
WEAPON_RE = re.compile(
    r"^(?:(?P<count>\d+)x\s+)?"
    r"(?P<wname>.+?)\s+"
    r"(?P<rng>\d+\"\s*|-)\s+"
    r"(?P<atk>A\d+)\s+"
    r"(?P<ap>-|\d+)\s+"
    r"(?P<spe>.+)$"
)
COST_RE = re.compile(r"^(?P<text>.+?)\s+(?P<cost>(?:\+\d+pts|Free))$")

def _is_group_header(line: str) -> bool:
    return bool(re.match(r"^(Upgrade|Replace|Any model)", line))

def _parse_attacks(atk: str) -> int:
    return int(atk.lstrip("A"))

def _parse_ap(ap: str) -> Optional[int]:
    if ap == "-":
        return None
    return int(ap)

def _split_rules(text: str) -> List[str]:
    text = text.replace(" ,", ",").strip().strip(",")
    return [p.strip() for p in text.split(",") if p.strip()]

def extract_lines(pdf_path: str) -> List[str]:
    try:
        import pdfplumber  # type: ignore
    except Exception as e:
        raise RuntimeError("Missing dependency pdfplumber. Install with: pip install pdfplumber") from e

    lines: List[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=1.5, y_tolerance=2, use_text_flow=True) or ""
            for raw in text.splitlines():
                line = raw.strip()
                if not line:
                    continue
                if line.startswith("GF - "):
                    continue
                if re.fullmatch(r"V\d+(\.\d+)*", line):
                    continue
                if re.fullmatch(r"\d+", line):
                    continue
                if re.fullmatch(r"\.{2,}", line):
                    continue
                lines.append(line)
    return lines

def parse_units(lines: List[str]) -> List[Dict[str, Any]]:
    units: List[Dict[str, Any]] = []
    i = 0
    n = len(lines)

    while i < n:
        m = UNIT_HEADER_RE.match(lines[i])
        if not m:
            i += 1
            continue

        unit: Dict[str, Any] = {
            "name": m.group("name").strip(),
            "size": int(m.group("size")),
            "base_points": int(m.group("pts")),
            "quality": None,
            "defense": None,
            "tough": None,
            "special_rules": [],
            "weapons": [],
            "options": [],
        }
        i += 1

        if i < n:
            ms = STATS_RE.match(lines[i])
            if ms:
                unit["quality"] = int(ms.group("qua"))
                unit["defense"] = int(ms.group("def"))
                if ms.group("tough"):
                    unit["tough"] = int(ms.group("tough"))
                i += 1

        rules_lines: List[str] = []
        while i < n and lines[i] != WEAPON_HEADER and not UNIT_HEADER_RE.match(lines[i]):
            rules_lines.append(lines[i])
            i += 1
        if rules_lines:
            unit["special_rules"] = _split_rules(" ".join(rules_lines))

        if i < n and lines[i] == WEAPON_HEADER:
            i += 1

        while i < n:
            line = lines[i]
            if UNIT_HEADER_RE.match(line) or _is_group_header(line) or line == WEAPON_HEADER:
                break
            wm = WEAPON_RE.match(line)
            if wm:
                spe = wm.group("spe").strip()
                weapon = {
                    "count": int(wm.group("count") or 1),
                    "name": wm.group("wname").strip(),
                    "range": wm.group("rng").replace(" ", ""),
                    "attacks": _parse_attacks(wm.group("atk")),
                    "ap": _parse_ap(wm.group("ap")),
                    "special": [] if spe == "-" else [s.strip() for s in spe.split(",") if s.strip() and s.strip() != "-"],
                }
                unit["weapons"].append(weapon)
            i += 1

        while i < n and not UNIT_HEADER_RE.match(lines[i]):
            if lines[i] == WEAPON_HEADER:
                i += 1
                continue

            if _is_group_header(lines[i]):
                header = lines[i]
                group = {"header": header, "options": []}
                i += 1

                while i < n and not UNIT_HEADER_RE.match(lines[i]) and not _is_group_header(lines[i]) and lines[i] != WEAPON_HEADER:
                    opt_line = lines[i].strip()
                    if not opt_line:
                        i += 1
                        continue
                    cm = COST_RE.match(opt_line)
                    if cm:
                        text_part = cm.group("text").strip()
                        cost_part = cm.group("cost")
                        pts = 0 if cost_part == "Free" else int(cost_part.lstrip("+").rstrip("pts"))
                        group["options"].append({"text": text_part, "pts": pts})
                    else:
                        group["options"].append({"text": opt_line, "pts": None})
                    i += 1

                unit["options"].append(group)
            else:
                i += 1

        units.append(unit)

    return units

# =========================================================
# XLSX rule policy loader
# =========================================================
def _norm_rule_key(s: str) -> str:
    return " ".join(str(s).strip().split()).lower()

def rule_base(rule: str) -> str:
    rule = " ".join(str(rule).strip().split())
    if "(" in rule:
        return rule.split("(", 1)[0].strip()
    return rule

def type_to_bucket(t: str) -> str:
    t = " ".join(str(t).strip().split())
    t = t.upper().replace(" ", "_")
    return f"<{t}>"

def load_rule_policy_xlsx(xlsx_path: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise RuntimeError("Missing dependency openpyxl. Install with: pip install openpyxl") from e

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, {}

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    rule_col = None
    type_col = None
    for i, h in enumerate(header):
        hl = h.lower()
        if hl in ("rule_name", "rule", "name"):
            rule_col = i
        if hl in ("type", "classification", "class"):
            type_col = i
    if rule_col is None:
        rule_col = 0
    if type_col is None:
        type_col = 1 if len(header) > 1 else 0

    exact_map: Dict[str, str] = {}
    base_map: Dict[str, str] = {}

    for r in rows[1:]:
        if not r or rule_col >= len(r):
            continue
        rule_val = r[rule_col]
        if rule_val is None:
            continue
        rule_str = str(rule_val).strip()
        if not rule_str:
            continue
        t_val = r[type_col] if type_col < len(r) else ""
        t_str = str(t_val).strip() if t_val is not None else ""
        if not t_str:
            t_str = "EXCLUDE"

        ek = _norm_rule_key(rule_str)
        exact_map[ek] = t_str
        bk = _norm_rule_key(rule_base(rule_str))
        base_map.setdefault(bk, t_str)

    return exact_map, base_map

def normalize_rules_for_signature(rules_in: List[str], exact_map: Dict[str, str], base_map: Dict[str, str]) -> Tuple[str, ...]:
    rules = [norm_ws(r) for r in rules_in if norm_ws(r)]
    out: List[str] = []
    for r in rules:
        rk_exact = _norm_rule_key(r)
        rk_base = _norm_rule_key(rule_base(r))
        t = exact_map.get(rk_exact) or base_map.get(rk_base)

        if t is None:
            out.append(r)
            continue

        if RULE_POLICY == "keep":
            out.append(r)
        elif RULE_POLICY == "exclude":
            continue
        elif RULE_POLICY == "bucket":
            out.append(type_to_bucket(t))
        else:
            raise ValueError(f"Unknown RULE_POLICY={RULE_POLICY}")

    out = sorted(set(out), key=lambda x: x.lower())
    return tuple(out)

# =========================================================
# Stage-1 signature helpers (no text parsing)
# =========================================================
_ATTACKS_RE = re.compile(r"^A(?P<a>\d+)$", re.IGNORECASE)
_AP_RE = re.compile(r"^AP\((?P<ap>-?\d+)\)$", re.IGNORECASE)
_RNG_QUOTE_RE = re.compile(r'^(?P<r>\d+)"$')

def _split_top_level_commas(s: str) -> List[str]:
    out: List[str] = []
    buf: List[str] = []
    depth = 0
    for ch in s:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth = max(0, depth - 1)
        if ch == "," and depth == 0:
            piece = "".join(buf).strip()
            if piece:
                out.append(piece)
            buf = []
        else:
            buf.append(ch)
    tail = "".join(buf).strip()
    if tail:
        out.append(tail)
    return out

def looks_like_weapon_profile(inside_parens: Optional[str]) -> bool:
    if not inside_parens:
        return False
    if re.search(r"\bA\d+\b", inside_parens):
        return True
    if re.search(r"\bAP\(\s*-?\d+\s*\)", inside_parens):
        return True
    return False

def split_name_and_parens(text: str) -> Tuple[str, Optional[str]]:
    t = text.strip()
    m = re.match(r"^(?P<name>.+?)\s*\((?P<inside>.+)\)\s*$", t)
    if m:
        return m.group("name").strip(), m.group("inside").strip()
    return t, None

def parse_count_prefix(text: str) -> Tuple[int, str]:
    t = text.strip()
    m = re.match(r"^(?P<n>\d+)\s*[x×]\s*(?P<rest>.+)$", t, flags=re.I)
    if m:
        return int(m.group("n")), m.group("rest").strip()
    return 1, t

def weapon_key_from_profile(profile: str, weapon_name: str, rng_fallback: Optional[int] = None) -> Tuple[str, Optional[int], int, Optional[int], Tuple[str, ...]]:
    """
    Parse profile like: 18", A4, AP(1), Rending
    Returns:
      key_str like: N=Sword|R=18|A=4|AP=1|T=Rending;Reliable
    """
    tokens = [t.strip() for t in _split_top_level_commas(profile)]
    rng: Optional[int] = rng_fallback
    attacks = 0
    ap: Optional[int] = None
    tags: List[str] = []

    for t in tokens:
        if not t:
            continue
        mq = _RNG_QUOTE_RE.match(t)
        if mq:
            rng = int(mq.group("r"))
            continue
        ma = _ATTACKS_RE.match(t)
        if ma:
            attacks = int(ma.group("a"))
            continue
        map_ = _AP_RE.match(t)
        if map_:
            ap = int(map_.group("ap"))
            continue
        tags.append(t)

    tags_sorted = tuple(sorted([x for x in tags if x], key=lambda x: x.lower()))
    # Include weapon name in the key to preserve original names
    key = f"N={weapon_name}|R={'' if rng is None else rng}|A={attacks}|AP={'' if ap is None else ap}"
    if tags_sorted:
        key += "|T=" + ";".join(tags_sorted)
    return key, rng, attacks, ap, tags_sorted

def build_stage1_signature(points: int, rules: Tuple[str, ...], weapon_multiset: Dict[str, int]) -> str:
    items = [(k, c) for k, c in weapon_multiset.items() if c > 0]
    items.sort(key=lambda kv: kv[0])
    weapons_key = "|".join([f"{k}*{c}" for k, c in items])

    parts: List[str] = []
    if INCLUDE_POINTS_IN_STAGE1_SIGNATURE:
        parts.append(f"PTS={points}")
    parts.append(f"RULES={','.join(rules)}")
    parts.append(f"W={weapons_key}")
    return "||".join(parts)

# =========================================================
# Variant generation with precomputed deltas
# =========================================================
_WORD_TO_INT = {"one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
                "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10}

def norm_name(s: str) -> str:
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    if s.endswith("s") and not s.endswith("ss"):
        s2 = s[:-1]
        if len(s2) > 2:
            s = s2
    return s

def weapon_occurrences(unit: Dict[str, Any], weapon_name: str) -> int:
    wn = norm_name(weapon_name)
    occ = 0
    for w in unit.get("weapons", []):
        if norm_name(w.get("name", "")) == wn:
            occ += int(w.get("count", 1) or 1)
    return occ

def guess_upgrade_multiplier(header: str, unit_size: int) -> int:
    h = header.lower()
    if "all models" in h:
        return unit_size
    return 1

def header_pick_limit(header: str) -> Optional[int]:
    h = header.lower()
    m = re.search(r"\b(up to|upto)\s+(?P<n>\d+)\b", h)
    if m:
        return int(m.group("n"))
    m = re.search(r"\b(up to|upto)\s+(?P<w>one|two|three|four|five|six|seven|eight|nine|ten)\b", h)
    if m:
        return _WORD_TO_INT[m.group("w")]
    m = re.search(r"\bwith\s+(?P<n>\d+)\b", h)
    if m:
        return int(m.group("n"))
    m = re.search(r"\bwith\s+(?P<w>one|two|three|four|five|six|seven|eight|nine|ten)\b", h)
    if m:
        return _WORD_TO_INT[m.group("w")]
    return None

TREAT_UPGRADE_WITH_NO_QUANTITY_AS_ANY_SUBSET = True
ALLOW_EMPTY_FOR_ONE_SELECTION_GROUPS = True
ALLOW_MIXED_REPLACEMENTS_FOR_UP_TO = True

@dataclass(frozen=True)
class Variant:
    pts_delta: int
    add_rules: Tuple[str, ...]                 # raw rules (un-normalized)
    weapon_delta: Tuple[Tuple[str, int], ...]  # key->delta count (now includes weapon name)

def _extract_rules_from_choice(choice_text: str) -> List[str]:
    name_part, inside = split_name_and_parens(choice_text)
    if inside is not None:
        # If parentheses do NOT look like weapon profile, treat as rule payload.
        if not looks_like_weapon_profile(inside):
            return [inside.strip()] if inside.strip() else []
        return []
    # No parentheses: treat whole thing as rule-ish
    t = name_part.strip()
    return [t] if t else []

def _compute_option_weapon_key(opt: Dict[str, Any]) -> Tuple[str, int, int, List[str]]:
    """
    Compute the weapon key for an option to enable deduplication.
    Returns: (weapon_key, pts, count, rules)
    """
    pts = int(opt.get("pts", 0) or 0)
    txt = str(opt.get("text", "")).strip()
    name_part, inside = split_name_and_parens(txt)
    c, item_name = parse_count_prefix(name_part)

    if inside and looks_like_weapon_profile(inside):
        add_key = weapon_key_from_profile(inside, item_name)[0]
    else:
        add_key = f"N={item_name}|R=|A=0|AP="

    rules = _extract_rules_from_choice(txt)
    return (add_key, pts, c, rules)

def _dedupe_options_by_weapon_key(opts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Deduplicate options that produce the same weapon key.
    Keeps the first occurrence (lowest pts if tied).
    """
    seen: Dict[str, Dict[str, Any]] = {}
    for o in opts:
        key, pts, count, rules = _compute_option_weapon_key(o)
        # Create a composite key including pts and rules for full deduplication
        composite_key = (key, pts, count, tuple(rules))
        if composite_key not in seen:
            seen[composite_key] = o
    return list(seen.values())

def build_base_weapon_multiset(unit: Dict[str, Any]) -> Tuple[Dict[str, int], Dict[str, str]]:
    """
    Returns:
      base_multiset: weapon_key -> count (now includes weapon name in key)
      name_to_key: normalized weapon name -> weapon_key (first match)
    """
    base: Dict[str, int] = {}
    name_to_key: Dict[str, str] = {}

    for w in unit.get("weapons", []) or []:
        name = str(w.get("name", "")).strip()
        count = int(w.get("count", 1) or 1)
        rng_raw = str(w.get("range", "")).strip()
        rng: Optional[int] = None
        if rng_raw and rng_raw not in ("-", "—"):
            m = _RNG_QUOTE_RE.match(rng_raw)
            if m:
                rng = int(m.group("r"))
        attacks = int(w.get("attacks", 0) or 0)
        ap = w.get("ap", None)
        tags = tuple(sorted([str(s).strip() for s in (w.get("special") or []) if str(s).strip() and str(s).strip() != "-"], key=lambda x: x.lower()))

        # Include weapon name in the key to preserve original names
        key = f"N={name}|R={'' if rng is None else rng}|A={attacks}|AP={'' if ap is None else ap}"
        if tags:
            key += "|T=" + ";".join(tags)

        base[key] = base.get(key, 0) + count
        nn = norm_name(name)
        name_to_key.setdefault(nn, key)

    return base, name_to_key

def group_variants(unit: Dict[str, Any], group: Dict[str, Any], name_to_key: Dict[str, str]) -> List[Variant]:
    header = group.get("header", "").strip()
    opts = group.get("options", []) or []
    unit_size = int(unit.get("size", 1) or 1)

    h = header.lower()
    is_upgrade = h.startswith("upgrade")
    is_replace = h.startswith("replace") or h.startswith("any model may replace") or " replace " in h
    multiplier = guess_upgrade_multiplier(header, unit_size)

    out: List[Variant] = []

    def make(pts: int, add_rules: List[str], weapon_delta: Dict[str, int]) -> Variant:
        # compress
        wd = tuple(sorted([(k, v) for k, v in weapon_delta.items() if v != 0], key=lambda kv: kv[0]))
        ar = tuple([r for r in (x.strip() for x in add_rules) if r])
        return Variant(pts_delta=pts, add_rules=ar, weapon_delta=wd)

    # UPGRADE groups (rules-only in current model)
    if is_upgrade and not h.startswith("replace"):
        max_pick = header_pick_limit(header)
        if max_pick is None:
            if TREAT_UPGRADE_WITH_NO_QUANTITY_AS_ANY_SUBSET:
                indices = range(len(opts))
                for r in range(0, len(opts) + 1):
                    for combo in itertools.combinations(indices, r):
                        add_rules: List[str] = []
                        pts = 0
                        for i in combo:
                            o = opts[i]
                            pts_each = int(o.get("pts", 0) or 0)
                            pts += pts_each * multiplier
                            add_rules.extend(_extract_rules_from_choice(str(o.get("text", ""))))
                        out.append(make(pts, add_rules, {}))
                return out
            else:
                max_pick = 1

        indices = range(len(opts))
        start_r = 0 if ALLOW_EMPTY_FOR_ONE_SELECTION_GROUPS else (1 if max_pick == 1 else 0)
        for r in range(start_r, min(max_pick, len(opts)) + 1):
            for combo in itertools.combinations(indices, r):
                add_rules: List[str] = []
                pts = 0
                for i in combo:
                    o = opts[i]
                    pts_each = int(o.get("pts", 0) or 0)
                    pts += pts_each * multiplier
                    add_rules.extend(_extract_rules_from_choice(str(o.get("text", ""))))
                out.append(make(pts, add_rules, {}))
        return out

    # REPLACE groups
    if is_replace:
        mode = "bundle"
        slots = 1
        target_name = header

        m = re.match(r"^any model may replace\s+(?P<rest>.+)$", h)
        if m:
            mode = "per_slot"
            slots = unit_size
            target_name = m.group("rest").strip()

        m = re.match(r"^replace all\s+(?P<weapon>.+)$", h)
        if m:
            mode = "bundle_all"
            target_name = m.group("weapon").strip()
            slots = max(1, weapon_occurrences(unit, target_name))

        m = re.match(r"^replace any\s+(?P<weapon>.+)$", h)
        if m:
            mode = "per_slot"
            target_name = m.group("weapon").strip()
            slots = max(1, weapon_occurrences(unit, target_name))

        m = re.match(r"^replace one\s+(?P<weapon>.+)$", h)
        if m:
            mode = "bundle"
            target_name = m.group("weapon").strip()
            slots = 1

        m = re.match(r"^replace\s+(?:up to|upto)\s+(?P<n>\d+|one|two|three|four|five)\s+(?P<weapon>.+)$", h)
        if m:
            nraw = m.group("n")
            n = int(nraw) if nraw.isdigit() else _WORD_TO_INT.get(nraw, 2)
            target_name = m.group("weapon").strip()
            if ALLOW_MIXED_REPLACEMENTS_FOR_UP_TO:
                mode = "per_slot"
                slots = n
            else:
                mode = "bundle"
                slots = n

        m = re.match(r"^replace\s+(?P<n>\d+)\s*[x×]\s+(?P<weapon>.+)$", h)
        if m:
            mode = "bundle"
            slots = int(m.group("n"))
            target_name = m.group("weapon").strip()

        m = re.match(r"^replace\s+(?P<weapon>.+)$", h)
        if m and target_name == header:
            target_name = m.group("weapon").strip()
            occ = weapon_occurrences(unit, target_name)
            slots = max(1, occ)
            mode = "bundle"

        target_key = name_to_key.get(norm_name(target_name), "")
        def none_var() -> Variant:
            return make(0, [], {})

        if not opts:
            return [none_var()]

        # OPTIMIZATION 2: Deduplicate options that produce the same weapon key
        deduped_opts = _dedupe_options_by_weapon_key(opts)

        if mode in ("bundle", "bundle_all"):
            out.append(none_var())
            for o in deduped_opts:
                pts = int(o.get("pts", 0) or 0)
                txt = str(o.get("text", "")).strip()
                name_part, inside = split_name_and_parens(txt)
                c, item_name = parse_count_prefix(name_part)

                weapon_delta: Dict[str, int] = {}
                if target_key:
                    weapon_delta[target_key] = weapon_delta.get(target_key, 0) - int(slots)

                # Use the actual weapon name from the option text instead of a placeholder
                add_key = None
                if inside and looks_like_weapon_profile(inside):
                    add_key = weapon_key_from_profile(inside, item_name)[0]
                else:
                    # For melee weapons without explicit profile, use the weapon name
                    # instead of a generic placeholder
                    add_key = f"N={item_name}|R=|A=0|AP="

                # OPTIMIZATION 3: Skip self-replacements (replacing weapon with itself)
                if add_key == target_key:
                    continue  # No actual change - skip this variant

                weapon_delta[add_key] = weapon_delta.get(add_key, 0) + int(c)

                # rules added (if any non-weapon payload)
                add_rules = _extract_rules_from_choice(txt)

                out.append(make(pts, add_rules, weapon_delta))
            return out

        if mode == "per_slot":
            # OPTIMIZATION 1: Use combinations_with_replacement instead of product
            # This reduces combinations since order doesn't matter
            # e.g., (plasma, melta, plasma) is same as (plasma, plasma, melta)
            choices = [None] + deduped_opts
            out.append(none_var())

            # Use combinations_with_replacement for symmetry reduction
            # This generates unordered selections with repetition
            for pick_combo in itertools.combinations_with_replacement(choices, slots):
                pts_delta = 0
                weapon_delta: Dict[str, int] = {}
                add_rules: List[str] = []
                has_self_replacement = False

                for pick in pick_combo:
                    if pick is None:
                        continue
                    pts_delta += int(pick.get("pts", 0) or 0)
                    txt = str(pick.get("text", "")).strip()
                    name_part, inside = split_name_and_parens(txt)
                    c, item_name = parse_count_prefix(name_part)

                    if target_key:
                        weapon_delta[target_key] = weapon_delta.get(target_key, 0) - 1

                    # Use the actual weapon name instead of a placeholder
                    if inside and looks_like_weapon_profile(inside):
                        add_key = weapon_key_from_profile(inside, item_name)[0]
                    else:
                        # For melee weapons without explicit profile, use the weapon name
                        add_key = f"N={item_name}|R=|A=0|AP="

                    # OPTIMIZATION 3: Track if this is a self-replacement
                    if add_key == target_key:
                        has_self_replacement = True

                    weapon_delta[add_key] = weapon_delta.get(add_key, 0) + int(c)
                    add_rules.extend(_extract_rules_from_choice(txt))

                # Skip variants that only contain self-replacements (no net change)
                # Check if weapon_delta results in no actual change
                net_change = {k: v for k, v in weapon_delta.items() if v != 0}
                if not net_change and has_self_replacement:
                    continue  # Skip - this is effectively a no-op

                out.append(make(pts_delta, add_rules, weapon_delta))
            return out

    # fallback: treat each option as rule
    out.append(make(0, [], {}))
    for o in opts:
        pts = int(o.get("pts", 0) or 0)
        txt = str(o.get("text", "")).strip()
        out.append(make(pts, _extract_rules_from_choice(txt), {}))
    return out

# =========================================================
# Mixed-radix indexing (deterministic combo_index lineage)
# =========================================================
def total_combinations(radices: List[int]) -> int:
    t = 1
    for r in radices:
        t *= r
    return t

def index_to_choice_indices(idx: int, radices: List[int]) -> List[int]:
    out = [0] * len(radices)
    for pos in range(len(radices) - 1, -1, -1):
        r = radices[pos]
        idx, rem = divmod(idx, r)
        out[pos] = rem
    return out

# =========================================================
# Stage-1 parallel worker (globals to avoid re-pickling per task)
# =========================================================
_G_UNIT: Dict[str, Any] = {}
_G_GROUP_VARS: List[List[Variant]] = []
_G_RADICES: List[int] = []
_G_BASE_PTS: int = 0
_G_BASE_RULES: List[str] = []
_G_BASE_W: Dict[str, int] = {}
_G_EXACT_MAP: Dict[str, str] = {}
_G_BASE_MAP: Dict[str, str] = {}

def _init_worker(unit: Dict[str, Any],
                 group_vars: List[List[Variant]],
                 base_pts: int,
                 base_rules: List[str],
                 base_weapon_multiset: Dict[str, int],
                 exact_map: Dict[str, str],
                 base_map: Dict[str, str]) -> None:
    global _G_UNIT, _G_GROUP_VARS, _G_RADICES, _G_BASE_PTS, _G_BASE_RULES, _G_BASE_W, _G_EXACT_MAP, _G_BASE_MAP
    _G_UNIT = unit
    _G_GROUP_VARS = group_vars
    _G_RADICES = [len(vs) for vs in group_vars]
    _G_BASE_PTS = base_pts
    _G_BASE_RULES = base_rules
    _G_BASE_W = base_weapon_multiset
    _G_EXACT_MAP = exact_map
    _G_BASE_MAP = base_map

def _header_for(points: int, rules_sig: Tuple[str, ...]) -> str:
    name = str(_G_UNIT.get("name", "")).strip()
    size = int(_G_UNIT.get("size", 1) or 1)
    q = int(_G_UNIT.get("quality", 0) or 0)
    d = int(_G_UNIT.get("defense", 0) or 0)
    rules_str = ", ".join(rules_sig)
    return f"{name} [{size}] Q{q}+ D{d}+ | {points}pts | {rules_str}"

def _worker_stage1_range(start_idx: int, end_idx: int) -> Dict[str, Any]:
    """
    Returns a dict:
      sig -> {count:int, rep_idx:int, rep_header:str, points:int}
    """
    out: Dict[str, Dict[str, Any]] = {}

    for combo_idx in range(start_idx, end_idx):
        # Build combo by indexing variants
        if _G_RADICES:
            choice_idxs = index_to_choice_indices(combo_idx, _G_RADICES)
        else:
            choice_idxs = []

        pts_delta = 0
        add_rules: List[str] = []
        weapon_delta_acc: Dict[str, int] = {}

        for g_i, v_i in enumerate(choice_idxs):
            v = _G_GROUP_VARS[g_i][v_i]
            pts_delta += int(v.pts_delta)
            add_rules.extend(list(v.add_rules))
            for k, dv in v.weapon_delta:
                weapon_delta_acc[k] = weapon_delta_acc.get(k, 0) + int(dv)

        points = int(_G_BASE_PTS + pts_delta)

        # Rules -> normalize via XLSX policy
        rules_raw = _G_BASE_RULES + add_rules
        rules_sig = normalize_rules_for_signature(rules_raw, _G_EXACT_MAP, _G_BASE_MAP)

        # Weapons multiset -> base + deltas
        w: Dict[str, int] = dict(_G_BASE_W)
        for k, dv in weapon_delta_acc.items():
            w[k] = w.get(k, 0) + dv
            if w[k] < 0:
                w[k] = 0

        sig = build_stage1_signature(points, rules_sig, w)

        if sig not in out:
            out[sig] = {"count": 1, "rep_idx": combo_idx, "rep_header": _header_for(points, rules_sig), "points": points}
        else:
            out[sig]["count"] += 1
            # deterministic representative: lowest combo index wins
            if combo_idx < out[sig]["rep_idx"]:
                out[sig]["rep_idx"] = combo_idx
                out[sig]["rep_header"] = _header_for(points, rules_sig)
                out[sig]["points"] = points

    return out

# =========================================================
# Stage-2 reduce (no weapon grouping - preserve original names)
# =========================================================
def parse_signature_parts(sig: str) -> List[str]:
    return [p for p in sig.split("||") if p.strip()]

def split_sig_kv(parts: List[str]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for p in parts:
        if "=" in p:
            k, v = p.split("=", 1)
            out[k.strip()] = v
    return out

_ITEM_END_RE = re.compile(r"^(?P<field>.+?)\*(?P<count>\d+)$")

def _range_bucket(rng: Optional[int]) -> str:
    if rng is None:
        return "Melee"
    for b in RANGE_BUCKETS:
        if rng <= b:
            return str(b)
    return RANGE_BUCKET_HIGH

def parse_stage1_W_items(w_value: str) -> List[Tuple[str, Optional[int], Optional[int], Tuple[str, ...], int, int]]:
    """
    Parse weapon items from signature.
    Returns list of: (name, rng, ap, tags, attacks, count)
    """
    if not w_value:
        return []
    tokens = w_value.split("|")
    cur: List[str] = []
    out: List[Tuple[str, Optional[int], Optional[int], Tuple[str, ...], int, int]] = []

    def flush(fields: List[str], count: int) -> None:
        d: Dict[str, str] = {}
        for f in fields:
            if "=" in f:
                k, v = f.split("=", 1)
                d[k.strip()] = v.strip()

        # Extract weapon name
        name = d.get("N", "Unknown Weapon")

        r_raw = d.get("R", "")
        rng: Optional[int] = None
        if r_raw:
            try:
                rng = int(r_raw)
            except Exception:
                rng = None

        a_raw = d.get("A", "")
        attacks = int(a_raw) if a_raw.isdigit() else 0

        ap_raw = d.get("AP", "")
        ap: Optional[int] = None
        if ap_raw:
            try:
                ap = int(ap_raw)
            except Exception:
                ap = None

        t_raw = d.get("T", "")
        tags: Tuple[str, ...] = tuple(sorted([t for t in (t_raw.split(";") if t_raw else []) if t], key=lambda x: x.lower()))
        out.append((name, rng, ap, tags, attacks, count))

    for t in tokens:
        m = _ITEM_END_RE.match(t)
        if m:
            field = m.group("field")
            count = int(m.group("count"))
            cur.append(field)
            flush(cur, count)
            cur = []
        else:
            cur.append(t)

    return out

# =========================================================
# Weapon Grouping System - Groups by matching rules (range, AP, tags)
# =========================================================
@dataclass
class WeaponGroup:
    """Represents a group of weapons with identical rules but potentially different attacks."""
    group_id: str
    rng: Optional[int]
    ap: Optional[int]
    tags: Tuple[str, ...]
    total_attacks: int
    total_count: int
    source_weapons: List[Tuple[str, int, int]]  # List of (name, attacks, count)

def _weapon_rules_key(rng: Optional[int], ap: Optional[int], tags: Tuple[str, ...]) -> str:
    """Create a grouping key based on range, AP, and special rules."""
    r_str = "" if rng is None else str(rng)
    ap_str = "" if ap is None else str(ap)
    tags_str = ";".join(tags)
    return f"R={r_str}|AP={ap_str}|T={tags_str}"

def group_weapons_by_rules(items: List[Tuple[str, Optional[int], Optional[int], Tuple[str, ...], int, int]]) -> Tuple[List[WeaponGroup], Dict[str, List[str]]]:
    """
    Group weapons that have identical rules (range, AP, special rules).
    Weapons can differ in attack values - attacks are summed.

    Returns:
        - List of WeaponGroup objects
        - Dict mapping group_id -> list of source weapon names
    """
    # Group by rules key
    groups_map: Dict[str, List[Tuple[str, Optional[int], Optional[int], Tuple[str, ...], int, int]]] = {}

    for item in items:
        name, rng, ap, tags, attacks, count = item
        key = _weapon_rules_key(rng, ap, tags)
        groups_map.setdefault(key, []).append(item)

    # Build WeaponGroup objects
    weapon_groups: List[WeaponGroup] = []
    lineage: Dict[str, List[str]] = {}

    # Sort keys for deterministic ordering (melee first, then by range, then by AP)
    def sort_group_key(key: str) -> Tuple[int, int, str]:
        parts = dict(p.split("=", 1) for p in key.split("|") if "=" in p)
        r = parts.get("R", "")
        ap = parts.get("AP", "")
        rng_val = -1 if r == "" else int(r)
        ap_val = 0 if ap == "" else int(ap)
        return (rng_val, ap_val, parts.get("T", ""))

    sorted_keys = sorted(groups_map.keys(), key=sort_group_key)

    for idx, key in enumerate(sorted_keys, start=1):
        group_items = groups_map[key]

        # Extract common properties from first item
        _, rng, ap, tags, _, _ = group_items[0]

        # Sum attacks and collect source weapons
        total_attacks = 0
        total_count = 0
        source_weapons: List[Tuple[str, int, int]] = []
        source_names: List[str] = []

        for name, _, _, _, attacks, count in group_items:
            total_attacks += attacks * count
            total_count += count
            source_weapons.append((name, attacks, count))
            source_names.append(name)

        # Generate group ID
        group_id = f"WG{idx:03d}"

        weapon_groups.append(WeaponGroup(
            group_id=group_id,
            rng=rng,
            ap=ap,
            tags=tags,
            total_attacks=total_attacks,
            total_count=total_count,
            source_weapons=source_weapons
        ))

        lineage[group_id] = source_names

    return weapon_groups, lineage

def condensed_weapons_key_from_stage1_signature(stage1_sig: str) -> str:
    """
    Build a weapons key that groups weapons by rules (range, AP, tags).
    Weapons with different attacks but same rules are grouped together.
    """
    kv = split_sig_kv(parse_signature_parts(stage1_sig))
    items = parse_stage1_W_items(kv.get("W", ""))

    weapon_groups, _ = group_weapons_by_rules(items)

    # Build condensed key from groups
    parts_out: List[str] = []
    for wg in weapon_groups:
        r_str = "" if wg.rng is None else str(wg.rng)
        ap_str = "" if wg.ap is None else str(wg.ap)
        tags_str = ";".join(wg.tags)
        parts_out.append(f"GID={wg.group_id},R={r_str},AP={ap_str},T={tags_str},A={wg.total_attacks},C={wg.total_count}")

    return " | ".join(parts_out)

def condensed_weapons_line(stage1_sig: str) -> str:
    """
    Generate human-readable weapons line in OPR format.
    Format: WeaponName (A#, AP(#), SpecialRules)

    Groups weapons by rules - weapons with same range/AP/tags are combined.
    """
    kv = split_sig_kv(parse_signature_parts(stage1_sig))
    items = parse_stage1_W_items(kv.get("W", ""))

    weapon_groups, lineage = group_weapons_by_rules(items)

    chunks: List[str] = []
    for wg in weapon_groups:
        # Determine display name
        if len(wg.source_weapons) == 1:
            # Single weapon - use its original name
            display_name = wg.source_weapons[0][0]
        else:
            # Multiple weapons grouped - use group ID with source names
            display_name = wg.group_id

        # Build stats in OPR format: (A#, AP(#), rules)
        inner: List[str] = [f"A{wg.total_attacks}"]
        if wg.ap is not None:
            inner.append(f"AP({wg.ap})")
        inner.extend(list(wg.tags))

        # Build range prefix for ranged weapons
        if wg.rng is not None:
            range_prefix = f'{wg.rng}" '
        else:
            range_prefix = ""

        # Format: [count]x [range]Name (stats)
        if wg.total_count > 1 and len(wg.source_weapons) == 1:
            chunks.append(f"{wg.total_count}x {range_prefix}{display_name} ({', '.join(inner)})")
        else:
            chunks.append(f"{range_prefix}{display_name} ({', '.join(inner)})")

    return ", ".join(chunks)

def get_weapon_lineage(stage1_sig: str) -> Dict[str, List[str]]:
    """
    Get the mapping of weapon group IDs to source weapon names.
    """
    kv = split_sig_kv(parse_signature_parts(stage1_sig))
    items = parse_stage1_W_items(kv.get("W", ""))
    _, lineage = group_weapons_by_rules(items)
    return lineage

def inject_sg_into_header(header: str, sg_id: str, meta: str) -> str:
    # Header is already in expected format; just inject after unit name
    m = re.match(r"^(?P<name>.+?)\s+\[(?P<size>\d+)\]\s+Q(?P<q>\d\+)\s+D(?P<d>\d\+)\s+\|\s+(?P<pts>\d+)\s*pts\s+\|\s*(?P<rules>.*)\s*$", header.strip())
    label = f"{sg_id} ({meta})"
    if not m:
        return f"{label} {header}".strip()
    name = m.group("name").strip()
    rest = header[len(name):]
    return f"{name} - {label}{rest}"

def stage2_reduce(stage1_groups: List[Dict[str, Any]],
                 out_txt: Path,
                 out_stage2_json: Path,
                 out_index_json: Path) -> None:
    super_map: Dict[str, List[Dict[str, Any]]] = {}

    for g in stage1_groups:
        stage1_sig = g.get("signature")
        if not isinstance(stage1_sig, str):
            continue

        kv = split_sig_kv([p for p in parse_signature_parts(stage1_sig) if not p.startswith("PTS=")])  # ignore points
        rules = kv.get("RULES", "")
        wcond = condensed_weapons_key_from_stage1_signature(stage1_sig)
        super_sig = f"RULES={rules}||W={wcond}"
        super_map.setdefault(super_sig, []).append(g)

    sorted_super_sigs = sorted(super_map.keys(), key=lambda s: sha1(s))

    out_supergroups: List[Dict[str, Any]] = []
    lineage_index: Dict[str, Dict[str, Any]] = {}
    txt_lines: List[str] = []

    for idx, super_sig in enumerate(sorted_super_sigs, start=1):
        sg_id = f"SG{idx:04d}"
        children = super_map[super_sig]
        rep_child = children[0]  # deterministic due to sorting below

        # compute totals
        child_group_ids = [str(c.get("group_id")) for c in children if c.get("group_id") is not None]
        child_count = len(children)
        members_total = sum(int(c.get("count", 0) or 0) for c in children)

        # pick rep within children: smallest rep_combo_index wins
        rep_child = min(children, key=lambda c: int((c.get("representative", {}) or {}).get("combo_index_0based", 10**18)))

        rep_header = ((rep_child.get("representative") or {}) or {}).get("header", "")
        rep_sig = str(rep_child.get("signature", ""))

        # Get weapon lineage for this supergroup
        weapon_lineage = get_weapon_lineage(rep_sig)

        out_supergroups.append({
            "sg_id": sg_id,
            "supergroup_hash": sha1(super_sig)[:10],
            "signature": super_sig,
            "unit": rep_child.get("unit"),
            "count_child_groups": child_count,
            "count_members": members_total,
            "child_group_ids": child_group_ids,
            "child_groups": [{
                "group_id": c.get("group_id"),
                "count": c.get("count"),
                "representative_combo_index_0based": ((c.get("representative") or {}) or {}).get("combo_index_0based"),
            } for c in children],
            "representative_child_group": {
                "group_id": rep_child.get("group_id"),
                "signature": rep_sig,
                "representative": rep_child.get("representative"),
                "condensed_weapons_line": condensed_weapons_line(rep_sig),
                "weapon_group_lineage": weapon_lineage,
            },
        })

        lineage_index[sg_id] = {
            "supergroup_hash": sha1(super_sig)[:10],
            "child_group_ids": child_group_ids,
            "members_total": members_total,
            "weapon_group_lineage": weapon_lineage,
            "child_groups": [{
                "group_id": c.get("group_id"),
                "count": c.get("count"),
                "representative_combo_index_0based": ((c.get("representative") or {}) or {}).get("combo_index_0based"),
            } for c in children],
        }

        if isinstance(rep_header, str) and rep_header.strip():
            meta = f"child_groups={child_count} members={members_total}"
            txt_lines.append(inject_sg_into_header(rep_header, sg_id, meta))
            txt_lines.append(condensed_weapons_line(rep_sig))
            if ADD_BLANK_LINE_BETWEEN_UNITS:
                txt_lines.append("")

    out_stage2_json.write_text(json.dumps({
        "settings": {
            "IGNORE_POINTS": True,
            "WEAPON_GROUPING": "by_rules",  # Group weapons with same range/AP/tags
            "WEAPON_GROUP_ID_FORMAT": "WG001..",  # Format for weapon group IDs
            "SG_ID_FORMAT": "SG0001..",
            "NOTE": "Weapons with identical rules (range, AP, special rules) are grouped. "
                    "Attacks are summed. weapon_group_lineage maps group IDs to source weapon names.",
        },
        "total_stage1_groups": len(stage1_groups),
        "total_supergroups": len(out_supergroups),
        "supergroups": out_supergroups,
    }, indent=2), encoding="utf-8")

    out_index_json.write_text(json.dumps({
        "total_supergroups": len(out_supergroups),
        "index": lineage_index
    }, indent=2), encoding="utf-8")

    out_txt.write_text("\n".join(txt_lines).rstrip() + "\n", encoding="utf-8")

# =========================================================
# Unit pipeline: inline stage-1 with within-unit parallel generation
# =========================================================
def stage1_reduce_inline_parallel(unit: Dict[str, Any],
                                 exact_map: Dict[str, str],
                                 base_map: Dict[str, str],
                                 limit: int = 0) -> Dict[str, Any]:
    base_pts = int(unit.get("base_points", 0) or 0)
    base_rules = [str(x).strip() for x in (unit.get("special_rules") or []) if str(x).strip()]

    base_weapon_multiset, name_to_key = build_base_weapon_multiset(unit)

    groups = unit.get("options", []) or []
    group_vars: List[List[Variant]] = [group_variants(unit, g, name_to_key) for g in groups]
    radices = [len(vs) for vs in group_vars]
    total = total_combinations(radices) if radices else 1
    if limit and total > limit:
        total = limit

    if total <= 1 or WORKERS_PER_UNIT <= 1:
        _init_worker(unit, group_vars, base_pts, base_rules, base_weapon_multiset, exact_map, base_map)
        partial = _worker_stage1_range(0, total)
        merged = partial
    else:
        n_tasks = min(TASKS_PER_UNIT, total)
        chunk = math.ceil(total / n_tasks)

        merged: Dict[str, Dict[str, Any]] = {}
        with ProcessPoolExecutor(
            max_workers=WORKERS_PER_UNIT,
            initializer=_init_worker,
            initargs=(unit, group_vars, base_pts, base_rules, base_weapon_multiset, exact_map, base_map)
        ) as ex:
            futures = []
            for t in range(n_tasks):
                s = t * chunk
                e = min(total, (t + 1) * chunk)
                if s >= e:
                    continue
                futures.append(ex.submit(_worker_stage1_range, s, e))

            for fut in as_completed(futures):
                part = fut.result()
                for sig, info in part.items():
                    if sig not in merged:
                        merged[sig] = dict(info)
                    else:
                        merged[sig]["count"] += int(info["count"])
                        # representative: lowest index wins
                        if int(info["rep_idx"]) < int(merged[sig]["rep_idx"]):
                            merged[sig]["rep_idx"] = int(info["rep_idx"])
                            merged[sig]["rep_header"] = info["rep_header"]
                            merged[sig]["points"] = info["points"]

    # Build stage-1 groups list
    out_groups: List[Dict[str, Any]] = []
    for sig in sorted(merged.keys(), key=lambda s: sha1(s)):
        info = merged[sig]
        group_id = sha1(sig)[:10]
        rep_idx = int(info["rep_idx"])
        rep = {
            "combo_index_0based": rep_idx,
            "combo_index_1based": rep_idx + 1,
            "header": info["rep_header"],
        }
        out_groups.append({
            "group_id": group_id,
            "signature": sig,
            "unit": unit.get("name"),
            "points": int(info.get("points", 0) or 0),
            "count": int(info["count"]),
            "rule_policy": RULE_POLICY,
            "representative": rep,
        })

    payload = {
        "unit": unit.get("name"),
        "total_combinations_processed": total,
        "total_groups": len(out_groups),
        "settings": {
            "RULE_POLICY": RULE_POLICY,
            "INCLUDE_POINTS_IN_STAGE1_SIGNATURE": INCLUDE_POINTS_IN_STAGE1_SIGNATURE,
            "WORKERS_PER_UNIT": WORKERS_PER_UNIT,
            "TASKS_PER_UNIT": TASKS_PER_UNIT,
            "PRESERVE_WEAPON_NAMES": True,
            "OPTIMIZATIONS": {
                "SYMMETRY_REDUCTION": True,  # combinations_with_replacement instead of product
                "OPTION_DEDUPLICATION": True,  # dedupe options with same weapon key
                "SKIP_SELF_REPLACEMENTS": True,  # skip replacing weapon with itself
            },
            "NOTE": "Lineage uses combo_index (mixed-radix) instead of storing every member record.",
        },
        "groups": out_groups,
    }
    return payload

# =========================================================
# Runner
# =========================================================
def run() -> None:
    pdf_path = Path(PDF_PATH).expanduser()
    xlsx_path = Path(RULES_XLSX_PATH).expanduser()
    out_dir = Path(OUTPUT_DIR).expanduser()

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Rules XLSX not found: {xlsx_path}")

    ensure_dir(out_dir)

    # Load rule policy ONCE
    exact_map, base_map = load_rule_policy_xlsx(xlsx_path)

    # Parse PDF
    lines = extract_lines(str(pdf_path))
    units = parse_units(lines)
    units_payload = {"faction": FACTION_NAME, "version": FACTION_VERSION, "units": units}
    units_json_path = out_dir / f"{safe_filename(FACTION_NAME)}_{FACTION_VERSION}_units.json"
    units_json_path.write_text(json.dumps(units_payload, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"[OK] Parsed {len(units)} units -> {units_json_path}")
    print(f"[INFO] WORKERS_PER_UNIT={WORKERS_PER_UNIT}, TASKS_PER_UNIT={TASKS_PER_UNIT}")
    print(f"[INFO] Stage-1 is INLINE (no loadouts txt parsing). WRITE_UNGROUPED_LOADOUTS_TXT={WRITE_UNGROUPED_LOADOUTS_TXT}")
    print(f"[INFO] Weapon names are preserved (no grouping by range bucket)")
    print(f"[INFO] Optimizations enabled: symmetry reduction, option deduplication, skip self-replacements")

    for u in units:
        unit_name = str(u.get("name", "")).strip()
        if not unit_name:
            continue

        unit_dir = out_dir / safe_filename(unit_name)
        ensure_dir(unit_dir)

        stage1_json_path = unit_dir / f"{safe_filename(unit_name)}.loadouts.reduced.json"
        final_txt_path  = unit_dir / f"{safe_filename(unit_name)}.final.txt"
        final_json_path = unit_dir / f"{safe_filename(unit_name)}.final.supergroups.json"
        final_idx_path  = unit_dir / f"{safe_filename(unit_name)}.final.lineage_index.json"

        print(f"\n=== {unit_name} ===")

        # Stage-1 inline
        stage1_payload = stage1_reduce_inline_parallel(u, exact_map, base_map, limit=MAX_LOADOUTS_PER_UNIT)
        stage1_json_path.write_text(json.dumps(stage1_payload, indent=2), encoding="utf-8")
        print(f"[OK] Stage-1 inline reduced: {stage1_payload.get('total_groups'):,} groups "
              f"(from {stage1_payload.get('total_combinations_processed'):,} combos) -> {stage1_json_path}")

        # Stage-2 final
        stage2_reduce(stage1_payload["groups"], final_txt_path, final_json_path, final_idx_path)
        print(f"[OK] Stage-2 final -> {final_txt_path}")

    print("\nDONE.")

if __name__ == "__main__":
    run()
